"""
PharmaCare Pro — Flask + PostgreSQL Backend  (Partition Edition)
============================================================
Drop this file in the SAME folder as index.html, app.js, styles.css.

Install & Run:
    pip install flask flask-cors
    pip install psycopg2-binary flask flask-cors flask-jwt-extended
    python app.py

Then open:  http://localhost:5000

═══════════════════════════════════════════════════════════════
  DATABASE PARTITION ARCHITECTURE
═══════════════════════════════════════════════════════════════

Single pharmacare.db file with internal partitioning:

  ┌─ pharmacare.db ─────────────────────────────────────────┐
  │                                                          │
  │  settings         ← global (shared)                     │
  │  categories       ← global (shared across all modes)    │
  │  dashboard_resets ← per partition                       │
  │                                                          │
  │  ── MEDICINE DB ─────────────────────────────────────── │
  │  products    partition: 'wholesale' | 'retail' | 'both' │
  │  stock_ins   partition: 'wholesale' | 'retail' | 'both' │
  │                                                          │
  │  ── SALES HISTORY DB ──────────────────────────────── │
  │  bills       bill_store_type: 'wholesale' | 'retail'    │
  │  bill_items  (child of bills)                           │
  │                                                          │
  │  ── CREDIT DBs ─────────────────────────────────────── │
  │  credits       partition: 'wholesale' | 'both'          │
  │                (Wholesale mode: retailers who owe WS)   │
  │  shop_credits  partition: 'retail'    | 'both'          │
  │                (Retail mode: what shop owes suppliers)  │
  │                                                          │
  └──────────────────────────────────────────────────────────┘

Partition key mapping
─────────────────────────────────────────────────────────────
  Store Type              │  Partition Key
  ────────────────────────┼──────────────
  Wholesale Pharma        │  'wholesale'
  Retail Pharmacy         │  'retail'
  Hospital Pharmacy       │  'retail'
  Medical Store           │  'retail'
  Ayurvedic Store         │  'retail'

  partition = 'both'  → seed / migrated data visible to ALL modes
  partition = 'wholesale' → only visible in Wholesale Pharma mode
  partition = 'retail'    → only visible in all retail modes
"""
from dotenv import load_dotenv
load_dotenv()  
from flask import Flask, request, jsonify, abort, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required,
    get_jwt_identity, verify_jwt_in_request
)
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2, psycopg2.extras, psycopg2.pool, uuid, os, re, smtplib, random, time, json
import concurrent.futures
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import date, timedelta, datetime
from functools import wraps

# ─────────────────────────────────────────────────────────────
# APP SETUP
# ─────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL environment variable is not set.\n"
        "  Local dev  → create a .env file with DATABASE_URL=postgresql://...\n"
        "  Render     → add DATABASE_URL in the Render dashboard Environment tab.\n"
        "  Use the Supabase SESSION POOLER URL (port 6543) — works on IPv4 networks."
    )
# Only print a redacted version so passwords never appear in logs
_safe_url = DATABASE_URL.split("@")[-1] if "@" in DATABASE_URL else "(set)"
print(f"DATABASE_URL: ...@{_safe_url}")
app = Flask(__name__)
CORS(app)

# ─────────────────────────────────────────────────────────────
# JWT CONFIGURATION
# ─────────────────────────────────────────────────────────────
import secrets
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', secrets.token_hex(32))
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=30)
jwt = JWTManager(app)

# ── JWT identity serialisation fix (PyJWT 2.x requires sub = string) ──────
# Encode dict identity as JSON string so PyJWT is happy
@jwt.user_identity_loader
def _jwt_identity_serialiser(identity):
    if isinstance(identity, dict):
        return json.dumps(identity, separators=(',', ':'))
    return str(identity)

# ─────────────────────────────────────────────────────────────
# EMAIL CONFIG (optional — set environment variables to enable)
# SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM
# ─────────────────────────────────────────────────────────────
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

# Resend API (preferred — set RESEND_API_KEY in .env)
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
RESEND_FROM    = os.environ.get('RESEND_FROM', 'PharmaCare Pro <onboarding@resend.dev>')

# ── Serve frontend static files ───────────────────────────────
@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    if filename.startswith('api/'):
        abort(404)
    return send_from_directory(BASE_DIR, filename)


# ─────────────────────────────────────────────────────────────
# PARTITION CONSTANTS & HELPERS
# ─────────────────────────────────────────────────────────────
WHOLESALE_TYPE  = 'Wholesale Pharma'
RETAIL_TYPES    = {'Retail Pharma', 'Retail Pharmacy', 'Hospital Pharmacy', 'Medical Store', 'Ayurvedic Store'}
PARTITION_BOTH  = 'both'
PARTITION_WS    = 'wholesale'
PARTITION_RT    = 'retail'

def _store_partition(store_type: str) -> str:
    """Map settings.store_type → partition key ('wholesale' | 'retail')."""
    return PARTITION_WS if (store_type or '').strip() == WHOLESALE_TYPE else PARTITION_RT

def _jwt_partition() -> str:
    """Read pharmacy_type from current JWT token and return partition key."""
    try:
        identity = _get_identity()
        if isinstance(identity, dict):
            return _store_partition(identity.get('pharmacy_type', 'Retail Pharmacy'))
        return PARTITION_RT
    except Exception:
        return PARTITION_RT

def _jwt_store_type() -> str:
    """Return the pharmacy type from the current JWT token."""
    try:
        identity = _get_identity()
        if isinstance(identity, dict):
            return identity.get('pharmacy_type', 'Retail Pharmacy')
        return 'Retail Pharmacy'
    except Exception:
        return 'Retail Pharmacy'

def _partition_where(partition: str) -> tuple:
    """Return (WHERE clause snippet, params) that selects partition='both' OR partition=current."""
    return "partition IN (%s, 'both')", [partition]


# ─────────────────────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────────────────────
class _DBConn:
    """Wraps psycopg2 connection to mimic sqlite3 conn.execute() interface."""
    def __init__(self, pg_conn):
        self._c = pg_conn

    def execute(self, sql, params=None):
        cur = self._c.cursor()
        cur.execute(sql, params or ())
        return cur

    def commit(self):
        self._c.commit()

    def close(self):
        self._c.close()

    def rollback(self):
        try: self._c.rollback()
        except Exception: pass


# ── Connection Pool ───────────────────────────────────────────────────────────
# min=4 pre-opens 4 connections at startup so the first requests don't pay
# the 200ms TCP handshake cost. max=20 supports parallel queries in get_state.
_pool = None

def _get_pool():
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(
            4, 20,          # min=4 warm connections, max=20
            DATABASE_URL,
            cursor_factory=psycopg2.extras.DictCursor
        )
    return _pool

def get_db():
    pg = _get_pool().getconn()
    return _DBConn(pg)


class _PooledDBConn(_DBConn):
    """Like _DBConn but returns connection to pool on close()."""
    def close(self):
        try:
            _get_pool().putconn(self._c)
        except Exception:
            try: self._c.close()
            except Exception: pass

# Override get_db to return pooled connections
def get_db():  # noqa: F811
    pg = _get_pool().getconn()
    return _PooledDBConn(pg)


def init_db():
    """
    Create all tables.  Also runs lightweight migrations (ADD COLUMN)
    so existing databases are upgraded automatically on first run.
    """
    conn = get_db()
    # ── Create all tables (PostgreSQL) ─────────────────────────────────────────
    for _stmt in [
        """CREATE TABLE IF NOT EXISTS users (
            id              TEXT PRIMARY KEY,
            full_name       TEXT NOT NULL,
            email           TEXT UNIQUE NOT NULL,
            phone           TEXT DEFAULT '',
            pharmacy_type   TEXT NOT NULL DEFAULT 'Retail Pharmacy',
            password_hash   TEXT NOT NULL,
            drug_license    TEXT UNIQUE NOT NULL,
            gstin           TEXT UNIQUE NOT NULL,
            created_at      TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS settings (
            id                  SERIAL PRIMARY KEY,
            store_name          TEXT DEFAULT 'My Pharmacy',
            store_type          TEXT DEFAULT 'Retail Pharmacy',
            address             TEXT DEFAULT '',
            phone               TEXT DEFAULT '',
            email               TEXT DEFAULT '',
            license_no          TEXT DEFAULT '',
            gstin               TEXT DEFAULT '',
            default_gst         FLOAT DEFAULT 12,
            currency            TEXT DEFAULT '₹',
            low_stock_threshold INTEGER DEFAULT 10,
            expiry_alert_days   INTEGER DEFAULT 90,
            supplier_name       TEXT DEFAULT '',
            wholesaler          TEXT DEFAULT '',
            owner_name          TEXT DEFAULT '',
            wholesaler_id       TEXT DEFAULT '',
            shop_name           TEXT DEFAULT '',
            retailer_owner      TEXT DEFAULT '',
            wholesale_upi_qr    TEXT DEFAULT '',
            retail_upi_qr       TEXT DEFAULT '',
            next_bill_no        INTEGER DEFAULT 1,
            user_id             TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS categories (
            id      TEXT PRIMARY KEY,
            name    TEXT NOT NULL,
            "desc"  TEXT DEFAULT '',
            user_id TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS dashboard_resets (
            store_type_key  TEXT NOT NULL,
            reset_date      TEXT NOT NULL,
            user_id         TEXT DEFAULT NULL,
            PRIMARY KEY (store_type_key, user_id)
        )""",
        """CREATE TABLE IF NOT EXISTS products (
            id         TEXT PRIMARY KEY,
            name       TEXT NOT NULL,
            category   TEXT REFERENCES categories(id),
            unit       TEXT DEFAULT 'Tablet',
            purchase   FLOAT DEFAULT 0,
            sale       FLOAT DEFAULT 0,
            gst        FLOAT DEFAULT 12,
            stock      INTEGER DEFAULT 0,
            min_stock  INTEGER DEFAULT 10,
            sku        TEXT DEFAULT '',
            expiry     TEXT DEFAULT '',
            brand      TEXT DEFAULT '',
            hsn        TEXT DEFAULT '',
            "desc"     TEXT DEFAULT '',
            partition  TEXT DEFAULT 'both',
            user_id    TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS stock_ins (
            id           TEXT PRIMARY KEY,
            date         TEXT NOT NULL,
            product_id   TEXT REFERENCES products(id),
            product_name TEXT DEFAULT '',
            qty          INTEGER DEFAULT 0,
            price        FLOAT DEFAULT 0,
            batch        TEXT DEFAULT '',
            expiry       TEXT DEFAULT '',
            supplier     TEXT DEFAULT '',
            invoice_no   TEXT DEFAULT '',
            notes        TEXT DEFAULT '',
            partition    TEXT DEFAULT 'both',
            user_id      TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS bills (
            id               TEXT PRIMARY KEY,
            bill_no          TEXT DEFAULT '',
            date             TEXT DEFAULT '',
            customer         TEXT DEFAULT '',
            phone            TEXT DEFAULT '',
            doctor           TEXT DEFAULT '',
            rx               TEXT DEFAULT '',
            payment_mode     TEXT DEFAULT 'Cash',
            notes            TEXT DEFAULT '',
            subtotal         FLOAT DEFAULT 0,
            total_discount   FLOAT DEFAULT 0,
            total_gst        FLOAT DEFAULT 0,
            round_off        FLOAT DEFAULT 0,
            grand_total      FLOAT DEFAULT 0,
            bill_store_type  TEXT DEFAULT 'retail',
            ws_supplier      TEXT DEFAULT '',
            ws_owner         TEXT DEFAULT '',
            ws_gstin         TEXT DEFAULT '',
            shop_name        TEXT DEFAULT '',
            shopkeeper_gstin TEXT DEFAULT '',
            rt_shop          TEXT DEFAULT '',
            rt_owner         TEXT DEFAULT '',
            rt_gstin         TEXT DEFAULT '',
            rt_license       TEXT DEFAULT '',
            rt_email         TEXT DEFAULT '',
            rt_phone         TEXT DEFAULT '',
            created_at       TIMESTAMPTZ DEFAULT NOW(),
            user_id          TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS bill_items (
            id         TEXT PRIMARY KEY,
            bill_id    TEXT REFERENCES bills(id) ON DELETE CASCADE,
            product_id TEXT DEFAULT '',
            name       TEXT DEFAULT '',
            category   TEXT DEFAULT '',
            unit       TEXT DEFAULT '',
            qty        FLOAT DEFAULT 0,
            unit_price FLOAT DEFAULT 0,
            discount   FLOAT DEFAULT 0,
            gst_rate   FLOAT DEFAULT 0,
            gst_amt    FLOAT DEFAULT 0,
            line_total FLOAT DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS credits (
            id               TEXT PRIMARY KEY,
            date             TEXT DEFAULT '',
            shop_name        TEXT DEFAULT '',
            shopkeeper_name  TEXT DEFAULT '',
            phone            TEXT DEFAULT '',
            for_item         TEXT DEFAULT '',
            amount           FLOAT DEFAULT 0,
            method           TEXT DEFAULT 'Cash',
            status           TEXT DEFAULT 'Pending',
            partition        TEXT DEFAULT 'wholesale',
            user_id          TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS shop_credits (
            id                  TEXT PRIMARY KEY,
            supplier_id         TEXT DEFAULT '',
            supplier_name       TEXT DEFAULT '',
            owner_name          TEXT DEFAULT '',
            total_purchase      FLOAT DEFAULT 0,
            paid                FLOAT DEFAULT 0,
            payment_mode        TEXT DEFAULT 'Cash',
            pending             FLOAT DEFAULT 0,
            last_purchase_date  TEXT DEFAULT '',
            bill_date           TEXT DEFAULT '',
            status              TEXT DEFAULT 'Pending',
            partition           TEXT DEFAULT 'retail',
            user_id             TEXT DEFAULT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS purchase_records (
            id               TEXT PRIMARY KEY,
            date             TEXT DEFAULT '',
            medicine_name    TEXT DEFAULT '',
            qty              FLOAT DEFAULT 0,
            qty_unit         TEXT DEFAULT 'Box',
            amount_paid      FLOAT DEFAULT 0,
            party_name       TEXT DEFAULT '',
            party_type       TEXT DEFAULT 'Supplier',
            order_no         TEXT DEFAULT '',
            expected_delivery TEXT DEFAULT '',
            delivery_status  TEXT DEFAULT 'Pending',
            notes            TEXT DEFAULT '',
            user_id          TEXT DEFAULT NULL
        )""",
    ]:
        conn.execute(_stmt)

    # ── Migrations: PostgreSQL ADD COLUMN IF NOT EXISTS (idempotent) ──────────
    pg_migrations = [
        ("products",         "partition TEXT DEFAULT 'both'"),
        ("stock_ins",        "partition TEXT DEFAULT 'both'"),
        ("credits",          "partition TEXT DEFAULT 'both'"),
        ("shop_credits",     "partition TEXT DEFAULT 'both'"),
        ("products",         "user_id TEXT DEFAULT NULL"),
        ("stock_ins",        "user_id TEXT DEFAULT NULL"),
        ("bills",            "user_id TEXT DEFAULT NULL"),
        ("credits",          "user_id TEXT DEFAULT NULL"),
        ("shop_credits",     "user_id TEXT DEFAULT NULL"),
        ("settings",         "user_id TEXT DEFAULT NULL"),
        ("dashboard_resets", "user_id TEXT DEFAULT NULL"),
        ("categories",       "user_id TEXT DEFAULT NULL"),
        # Stores purchase cost at billing time — enables accurate profit without
        # needing to look up the current (possibly changed) product purchase price
        ("bill_items",       "purchase_price FLOAT DEFAULT 0"),
        # ── Unit/pack size fields (strip/box logic) ──────────────────────────
        ("products",         "pieces_per_strip INTEGER DEFAULT 10"),
        ("products",         "strips_per_box INTEGER DEFAULT 10"),
        ("products",         "purchase_unit TEXT DEFAULT 'strip'"),
        # bill_items: store what unit was sold in + pieces normalised
        ("bill_items",       "unit_type TEXT DEFAULT 'strip'"),
        ("bill_items",       "display_qty FLOAT DEFAULT 0"),
        ("bill_items",       "qty_in_pieces FLOAT DEFAULT 0"),
        # Wholesale: selling price per box set by the wholesaler (separate from MRP)
        ("products",         "selling_price FLOAT DEFAULT 0"),
        # bill_items: store amount before GST/discount for wholesale billing clarity
        ("bill_items",       "amount_before_tax FLOAT DEFAULT 0"),
        # bill_items: store MRP, selling price per box, and pack config at billing time
        ("bill_items",       "mrp_per_box FLOAT DEFAULT 0"),
        ("bill_items",       "selling_price_per_box FLOAT DEFAULT 0"),
        ("bill_items",       "strips_per_box INTEGER DEFAULT 10"),
        ("bill_items",       "pieces_per_strip INTEGER DEFAULT 10"),
        # Settings columns — idempotent safety net for any DB upgraded from older schema
        ("settings",         "wholesale_upi_qr TEXT DEFAULT ''"),
        ("settings",         "retail_upi_qr TEXT DEFAULT ''"),
        ("settings",         "wholesaler_id TEXT DEFAULT ''"),
        ("settings",         "shop_name TEXT DEFAULT ''"),
        ("settings",         "retailer_owner TEXT DEFAULT ''"),
        ("settings",         "wholesaler TEXT DEFAULT ''"),
        ("settings",         "owner_name TEXT DEFAULT ''"),
    ]
    for table, col_def in pg_migrations:
        col_name = col_def.split()[0]
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col_def}")
        except Exception:
            pass

    conn.execute("INSERT INTO settings (id) VALUES (1) ON CONFLICT DO NOTHING")

    # ── Fix FK constraints: ON DELETE SET NULL prevents ForeignKeyViolation ───
    # Deleting a category must NOT block if products still reference it.
    # Deleting a product must NOT block if stock_ins still reference it.
    for drop_sql, add_sql in [
        (
            "ALTER TABLE products DROP CONSTRAINT IF EXISTS products_category_fkey",
            "ALTER TABLE products ADD CONSTRAINT products_category_fkey "
            "FOREIGN KEY (category) REFERENCES categories(id) ON DELETE SET NULL"
        ),
        (
            "ALTER TABLE stock_ins DROP CONSTRAINT IF EXISTS stock_ins_product_id_fkey",
            "ALTER TABLE stock_ins ADD CONSTRAINT stock_ins_product_id_fkey "
            "FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE SET NULL"
        ),
    ]:
        try:
            conn.execute(drop_sql)
            conn.execute(add_sql)
        except Exception:
            pass

    conn.commit()
    conn.close()


# ─────────────────────────────────────────────────────────────
# OTP STORE (in-memory, expires in 10 minutes)
# ─────────────────────────────────────────────────────────────
_OTP_STORE = {}   # { user_id: { otp, new_email, expires_at } }

def _gen_otp():
    return str(random.randint(100000, 999999))

# ─────────────────────────────────────────────────────────────
# EMAIL HELPERS  (Resend → SMTP → terminal fallback)
# ─────────────────────────────────────────────────────────────
import urllib.request as _ul_req

def _send_email_resend(to: str, subject: str, html: str) -> bool:
    if not RESEND_API_KEY:
        return False
    try:
        payload = json.dumps({"from": RESEND_FROM, "to": [to], "subject": subject, "html": html}).encode()
        req = _ul_req.Request(
            "https://api.resend.com/emails", data=payload,
            headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
            method="POST")
        with _ul_req.urlopen(req, timeout=10) as r:
            result = json.loads(r.read())
        print(f"  ✓ Email via Resend -> {to}  id={result.get('id','')}")
        return True
    except Exception as e:
        print(f"  x Resend failed -> {e}")
        return False

def _send_email_smtp(to: str, subject: str, html: str) -> bool:
    if not SMTP_USER or not SMTP_PASS:
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = SMTP_FROM or SMTP_USER
        msg["To"]      = to
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.ehlo(); s.starttls(); s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_FROM or SMTP_USER, to, msg.as_string())
        print(f"  ✓ Email via SMTP -> {to}")
        return True
    except Exception as e:
        print(f"  x SMTP failed -> {e}")
        return False

def _send_email(to: str, subject: str, html: str) -> bool:
    """Try Resend then SMTP. Never raises."""
    return _send_email_resend(to, subject, html) or _send_email_smtp(to, subject, html)

def _send_welcome_email(to_email: str, full_name: str, pharmacy_type: str):
    html = f"""<html><body style="font-family:sans-serif;max-width:600px;margin:0 auto;padding:20px">
    <div style="background:linear-gradient(135deg,#1e40af,#0891b2);padding:30px;border-radius:12px 12px 0 0;text-align:center">
      <h1 style="color:white;margin:0;font-size:28px">PharmaCare Pro</h1>
    </div>
    <div style="background:#f8fafc;padding:30px;border-radius:0 0 12px 12px;border:1px solid #e2e8f0">
      <h2 style="color:#0f172a">Welcome, {full_name}!</h2>
      <p style="color:#475569">Your <strong>{pharmacy_type}</strong> account is ready on PharmaCare Pro.</p>
      <p style="color:#64748b;font-size:14px">Email: {to_email}</p>
    </div></body></html>"""
    if not _send_email(to_email, "Welcome to PharmaCare Pro!", html):
        print(f"  i  Welcome email skipped (no email provider configured) -> {to_email}")


# ─────────────────────────────────────────────────────────────
# VALIDATION HELPERS
# ─────────────────────────────────────────────────────────────
DRUG_LICENSE_RE = re.compile(
    r'^(DL|dl)[- ]%s[A-Za-z0-9\-/]{4,25}$|'         # DL-XX-XXXXXX style
    r'^[A-Za-z]{2}[/ ]%s\d{2}[/ ]%s\d{5}$|'          # MH/20/12345 style
    r'^[A-Za-z0-9/\- ]{5,30}$'                      # catch-all alphanumeric
)
GSTIN_RE = re.compile(
    r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
)

def _validate_drug_license(dl: str) -> bool:
    dl = (dl or '').strip().upper()
    return bool(dl and len(dl) >= 5)

def _validate_gstin(gstin: str) -> bool:
    gstin = (gstin or '').strip().upper()
    return bool(GSTIN_RE.match(gstin))

def _validate_password(pw: str) -> str:
    """Returns error message or empty string if valid."""
    if len(pw) < 8:
        return "Password must be at least 8 characters"
    if not re.search(r'[A-Za-z]', pw):
        return "Password must contain at least one letter"
    if not re.search(r'[0-9]', pw):
        return "Password must contain at least one number"
    return ""


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def uid():
    return uuid.uuid4().hex[:12]

def _copy_seed_products_to_user(conn, user_id: str):
    """
    Copy seed products (user_id IS NULL) into a specific user's account.
    Uses count comparison as fast-path: if user already has ≥ seed count, skip entirely.
    Falls back to bulk SELECT diff if some seeds are missing.
    """
    # ── FAST PATH: 2 COUNT queries — skip entirely if already done ────────────
    seed_count = conn.execute(
        "SELECT COUNT(*) FROM products WHERE user_id IS NULL"
    ).fetchone()[0]
    if seed_count == 0:
        return  # Nothing to copy
    user_count = conn.execute(
        "SELECT COUNT(*) FROM products WHERE user_id=%s", (user_id,)
    ).fetchone()[0]
    if user_count >= seed_count:
        # User already has all seeds — nothing to do
        return
    # ─────────────────────────────────────────────────────────────────────────

    # ── Categories: bulk check — 1 query instead of N ────────────────────────
    seed_cats = conn.execute(
        "SELECT * FROM categories WHERE user_id IS NULL"
    ).fetchall()
    existing_cats = {
        r['id'] for r in conn.execute(
            "SELECT id FROM categories WHERE user_id=%s", (user_id,)
        ).fetchall()
    }
    for cat in seed_cats:
        if cat['id'] not in existing_cats:
            conn.execute(
                'INSERT INTO categories(id,name,"desc",user_id) VALUES(%s,%s,%s,%s) ON CONFLICT DO NOTHING',
                (cat['id'], cat['name'], cat['desc'], user_id)
            )

    # ── Products: ONE bulk query to get existing, diff in Python ─────────────
    seed_prods = conn.execute(
        "SELECT * FROM products WHERE user_id IS NULL"
    ).fetchall()
    existing_set = {
        (r['name'], r['partition'])
        for r in conn.execute(
            "SELECT name, partition FROM products WHERE user_id=%s", (user_id,)
        ).fetchall()
    }
    inserted = 0
    for p in seed_prods:
        if (p['name'], p['partition']) in existing_set:
            continue
        new_id = uid()
        conn.execute("""
            INSERT INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,
               sku,expiry,brand,hsn,"desc",partition,user_id)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (new_id, p['name'], p['category'], p['unit'],
              p['purchase'], p['sale'], p['gst'], p['stock'], p['min_stock'],
              p['sku'], p['expiry'], p['brand'], p['hsn'], p['desc'],
              p['partition'], user_id))
        inserted += 1

    conn.commit()
    if inserted:
        print(f"  ✓ Seed copy done for {user_id[:8]}: {inserted} new products added")

def today_str():
    return date.today().isoformat()

def row(r):
    return dict(r) if r else None

def rows(rs):
    return [dict(r) for r in rs]

def _get_identity() -> dict:
    """Deserialise JWT identity. Always returns a dict with at least user_id."""
    raw = get_jwt_identity()  # intentional — this IS _get_identity's internals
    if isinstance(raw, dict):
        return raw          # older tokens stored as dict (backwards compat)
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except Exception:
            return {'user_id': raw}
    return {}

def require_json(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if request.method in ('POST', 'PUT', 'PATCH') and not request.is_json:
            return jsonify({"error": "Content-Type must be application/json"}), 400
        return f(*a, **kw)
    return wrapper

def expiry_days_left(exp_month):
    if not exp_month:
        return 9999
    try:
        y, m = int(exp_month[:4]), int(exp_month[5:7])
        last = date(y, m, 1) + timedelta(days=32)
        last = last.replace(day=1) - timedelta(days=1)
        return (last - date.today()).days
    except Exception:
        return 9999

def _settings_out(s):
    return {
        "storeName":         s.get("store_name",          "My Pharmacy"),
        "storeType":         s.get("store_type",          "Retail Pharmacy"),
        "address":           s.get("address",             ""),
        "phone":             s.get("phone",               ""),
        "email":             s.get("email",               ""),
        "license":           s.get("license_no",          ""),
        "gstin":             s.get("gstin",               ""),
        "defaultGst":        s.get("default_gst",         12),
        "currency":          s.get("currency",            "₹"),
        "lowStockThreshold": s.get("low_stock_threshold", 10),
        "expiryAlertDays":   s.get("expiry_alert_days",   90),
        "wholesaler":        s.get("wholesaler",          ""),
        "ownerName":         s.get("owner_name",          ""),
        "wholesalerId":      s.get("wholesaler_id",       ""),
        "shopName":          s.get("shop_name",           ""),
        "retailerOwner":     s.get("retailer_owner",      ""),
        "wholesaleUpiQr":    s.get("wholesale_upi_qr",    ""),
        "retailUpiQr":       s.get("retail_upi_qr",       ""),
        "nextBillNo":        s.get("next_bill_no",        1),
    }

def _product_out(p):
    return {
        "id":            p["id"],
        "name":          p["name"],
        "category":      p["category"] or "",
        "unit":          p["unit"]     or "Tablet",
        "purchase":      p["purchase"] or 0,
        "sale":          p["sale"]     or 0,
        "gst":           p["gst"]      or 0,
        "stock":         p["stock"]    or 0,
        "minStock":      p["min_stock"] or 10,
        "sku":           p["sku"]      or "",
        "expiry":        p["expiry"]   or "",
        "brand":         p["brand"]    or "",
        "hsn":           p["hsn"]      or "",
        "desc":          p["desc"]     or "",
        "partition":     p["partition"] if "partition" in p.keys() else PARTITION_BOTH,
        # Pack-size / unit logic fields
        "piecesPerStrip": int(p["pieces_per_strip"]) if "pieces_per_strip" in p.keys() and p["pieces_per_strip"] else 10,
        "stripsPerBox":   int(p["strips_per_box"])   if "strips_per_box"   in p.keys() and p["strips_per_box"]   else 10,
        "purchaseUnit":   p["purchase_unit"] if "purchase_unit" in p.keys() and p["purchase_unit"] else "strip",
        # Wholesale: selling price per box set by the wholesaler (distinct from MRP)
        "sellingPrice":   float(p["selling_price"]) if "selling_price" in p.keys() and p["selling_price"] else 0,
    }

def _bill_item_out(i):
    return {
        "id":            i["id"],
        "productId":     i["product_id"] or "",
        "name":          i["name"]       or "",
        "category":      i["category"]   or "",
        "unit":          i["unit"]       or "",
        "qty":           i["qty"]        or 0,
        "unitPrice":     i["unit_price"] or 0,
        "discount":      i["discount"]   or 0,
        "gstRate":       i["gst_rate"]   or 0,
        "gstAmt":        i["gst_amt"]    or 0,
        "lineTotal":     i["line_total"] or 0,
        # Billing-time purchase price — accurate even if product is later edited/deleted
        "purchasePrice": float(i["purchase_price"]) if "purchase_price" in i.keys() and i["purchase_price"] else 0,
        # Unit-type fields
        "unitType":          i["unit_type"]          if "unit_type"          in i.keys() and i["unit_type"]          else "strip",
        "displayQty":        i["display_qty"]         if "display_qty"         in i.keys() and i["display_qty"]         else (i["qty"] or 0),
        "qtyInPieces":       i["qty_in_pieces"]       if "qty_in_pieces"       in i.keys() and i["qty_in_pieces"]       else (i["qty"] or 0),
        # Wholesale billing fields needed for profit graph and bill preview
        "amountBeforeTax":   float(i["amount_before_tax"])    if "amount_before_tax"    in i.keys() and i["amount_before_tax"]    else 0,
        "mrpPerBox":         float(i["mrp_per_box"])           if "mrp_per_box"           in i.keys() and i["mrp_per_box"]           else 0,
        "sellingPricePerBox":float(i["selling_price_per_box"]) if "selling_price_per_box" in i.keys() and i["selling_price_per_box"] else 0,
        "stripsPerBox":      int(i["strips_per_box"])          if "strips_per_box"        in i.keys() and i["strips_per_box"]        else 10,
        "piecesPerStrip":    int(i["pieces_per_strip"])        if "pieces_per_strip"      in i.keys() and i["pieces_per_strip"]      else 10,
    }

def _bill_out(b, conn):
    items = conn.execute(
        "SELECT * FROM bill_items WHERE bill_id=%s", (b["id"],)
    ).fetchall()
    return {
        "id":              b["id"],
        "billNo":          b["bill_no"]          or "",
        "date":            b["date"]             or "",
        "customer":        b["customer"]         or "",
        "phone":           b["phone"]            or "",
        "doctor":          b["doctor"]           or "",
        "rx":              b["rx"]               or "",
        "paymentMode":     b["payment_mode"]     or "Cash",
        "notes":           b["notes"]            or "",
        "subtotal":        b["subtotal"]         or 0,
        "totalDiscount":   b["total_discount"]   or 0,
        "totalGst":        b["total_gst"]        or 0,
        "roundOff":        b["round_off"]        or 0,
        "grandTotal":      b["grand_total"]      or 0,
        "billStoreType":   b["bill_store_type"]  or "retail",
        "wsSupplier":      b["ws_supplier"]      or "",
        "wsOwner":         b["ws_owner"]         or "",
        "wsGstin":         b["ws_gstin"]         or "",
        "shopName":        b["shop_name"]        or "",
        "shopkeeperGstin": b["shopkeeper_gstin"] or "",
        "rtShop":          b["rt_shop"]          or "",
        "rtOwner":         b["rt_owner"]         or "",
        "rtGstin":         b["rt_gstin"]         or "",
        "rtLicense":       b["rt_license"]       or "",
        "rtEmail":         b["rt_email"]         or "",
        "rtPhone":         b["rt_phone"]         or "",
        "items": [_bill_item_out(i) for i in items],
    }

def _credit_out(c):
    return {
        "id":             c["id"],
        "date":           c["date"]            or "",
        "shopName":       c["shop_name"]       or "",
        "shopkeeperName": c["shopkeeper_name"] or "",
        "phone":          c["phone"]           or "",
        "forItem":        c["for_item"]        or "",
        "amount":         c["amount"]          or 0,
        "method":         c["method"]          or "Cash",
        "status":         c["status"]          or "Pending",
        "partition":      c["partition"] if "partition" in c.keys() else PARTITION_WS,
    }

def _shop_credit_out(s):
    return {
        "id":               s["id"],
        "supplierId":       s["supplier_id"]        or "",
        "supplierName":     s["supplier_name"]      or "",
        "ownerName":        s["owner_name"]         or "",
        "totalPurchase":    s["total_purchase"]     or 0,
        "paid":             s["paid"]               or 0,
        "paymentMode":      s["payment_mode"]       or "Cash",
        "pending":          s["pending"]            or 0,
        "lastPurchaseDate": s["last_purchase_date"] or "",
        "billDate":         s["bill_date"]          or "",
        "status":           s["status"]             or "Pending",
        "partition":        s["partition"] if "partition" in s.keys() else PARTITION_RT,
    }

def _calc_item(unit_price, qty, discount, gst_rate):
    """Returns (gst_amt, line_total) for one bill line."""
    line_gross = qty * unit_price
    disc_amt   = line_gross * discount / 100
    taxable    = line_gross - disc_amt
    gst_amt    = taxable * (gst_rate / 100)
    line_total = taxable + gst_amt
    return round(gst_amt, 2), round(line_total, 2)

def _calc_totals(items):
    subtotal   = sum(it["qty"] * it["unit_price"] for it in items)
    total_disc = sum(it["qty"] * it["unit_price"] * it["discount"] / 100 for it in items)
    total_gst  = sum(it["gst_amt"] for it in items)
    raw        = subtotal - total_disc + total_gst
    grand      = round(raw)
    return {
        "subtotal":       round(subtotal,   2),
        "total_discount": round(total_disc, 2),
        "total_gst":      round(total_gst,  2),
        "round_off":      round(grand - raw, 2),
        "grand_total":    grand,
    }

def _get_settings(conn):
    r = conn.execute("SELECT * FROM settings ORDER BY id LIMIT 1").fetchone()
    return dict(r) if r else {}

def _get_user_by_id(conn, user_id: str):
    r = conn.execute("SELECT * FROM users WHERE id=%s", (user_id,)).fetchone()
    return dict(r) if r else {}

def _settings_out_with_user(s, user: dict = None):
    base = {
        "storeName":         s.get("store_name",          "My Pharmacy"),
        "storeType":         user.get("pharmacy_type", s.get("store_type", "Retail Pharmacy")) if user else s.get("store_type", "Retail Pharmacy"),
        "address":           s.get("address",             ""),
        "phone":             user.get("phone", s.get("phone", "")) if user else s.get("phone", ""),
        "email":             user.get("email", s.get("email", "")) if user else s.get("email", ""),
        "license":           user.get("drug_license", s.get("license_no", "")) if user else s.get("license_no", ""),
        "gstin":             user.get("gstin", s.get("gstin", "")) if user else s.get("gstin", ""),
        "defaultGst":        s.get("default_gst",         12),
        "currency":          s.get("currency",            "₹"),
        "lowStockThreshold": s.get("low_stock_threshold", 10),
        "expiryAlertDays":   s.get("expiry_alert_days",   90),
        "wholesaler":        s.get("wholesaler",          ""),
        "ownerName":         user.get("full_name", s.get("owner_name", "")) if user else s.get("owner_name", ""),
        "wholesalerId":      s.get("wholesaler_id",       ""),
        "shopName":          s.get("shop_name",           ""),
        "retailerOwner":     s.get("retailer_owner",      ""),
        "wholesaleUpiQr":    s.get("wholesale_upi_qr",    ""),
        "retailUpiQr":       s.get("retail_upi_qr",       ""),
        "nextBillNo":        s.get("next_bill_no",        1),
        # Locked user fields
        "userName":          user.get("full_name", "") if user else "",
        "userEmail":         user.get("email", "") if user else "",
        "userPhone":         user.get("phone", "") if user else "",
        "pharmacyTypeLocked": user.get("pharmacy_type", "") if user else "",
        "drugLicenseLocked": user.get("drug_license", "") if user else "",
        "gstinLocked":       user.get("gstin", "") if user else "",
    }
    return base

def _next_bill_no(conn):
    s = _get_settings(conn)
    n = s.get("next_bill_no", 1)
    return n, str(n).zfill(4)

def _bill_type_filter(partition: str) -> str:
    """Map partition key → bill_store_type filter expression."""
    # For bills we already have bill_store_type.
    # Wholesale sees only 'wholesale', retail sees everything that is NOT 'wholesale'.
    if partition == PARTITION_WS:
        return "bill_store_type = 'wholesale'"
    return "bill_store_type != 'wholesale'"


# ═════════════════════════════════════════════════════════════
# API ROUTES
# ═════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════
# AUTH ROUTES
# ═════════════════════════════════════════════════════════════

VALID_PHARMACY_TYPES = {'Retail Pharma', 'Wholesale Pharma',
    # legacy values accepted for existing accounts:
    'Retail Pharmacy', 'Hospital Pharmacy', 'Medical Store', 'Ayurvedic Store'}

@app.route('/api/auth/register', methods=['POST'])
@require_json
def auth_register():
    d             = request.get_json()
    email         = (d.get('email') or '').strip().lower()
    phone         = (d.get('phone') or '').strip()
    pharmacy_type = (d.get('pharmacyType') or '').strip()
    password      = d.get('password') or ''
    confirm_pw    = d.get('confirmPassword') or ''
    drug_license  = (d.get('drugLicense') or '').strip().upper()
    gstin         = (d.get('gstin') or '').strip().upper()
    address       = (d.get('address') or '').strip()
    default_gst   = float(d.get('defaultGst') or 12)
    low_stock     = int(d.get('lowStockThreshold') or 10)
    expiry_days   = int(d.get('expiryAlertDays') or 90)

    # Type-specific owner / business fields
    is_wholesale  = pharmacy_type == 'Wholesale Pharma'
    owner_name    = (d.get('ownerName') or '').strip()       # wholesale owner
    wholesaler    = (d.get('wholesaler') or '').strip()       # wholesale business name
    wholesaler_id = (d.get('wholesalerId') or '').strip()     # wholesale ID
    shop_name     = (d.get('shopName') or '').strip()         # retail shop name
    retailer_owner= (d.get('retailerOwner') or '').strip()    # retail owner name

    # Derive full_name from owner field (no separate full-name input)
    full_name = owner_name if is_wholesale else retailer_owner

    # Validation
    if not full_name:
        err = "Wholesaler Owner Name is required" if is_wholesale else "Retailer / Owner Name is required"
        return jsonify({"error": err}), 400
    if not email or '@' not in email:
        return jsonify({"error": "Valid email is required"}), 400
    if not phone or len(phone) < 10:
        return jsonify({"error": "Valid 10-digit phone number is required"}), 400
    if pharmacy_type not in VALID_PHARMACY_TYPES:
        return jsonify({"error": "Please select Retail Pharma or Wholesale Pharma"}), 400
    if is_wholesale and not wholesaler:
        return jsonify({"error": "Wholesaler (Business Name) is required"}), 400
    if not is_wholesale and not shop_name:
        return jsonify({"error": "Retail / Shop Name is required"}), 400
    pw_err = _validate_password(password)
    if pw_err:
        return jsonify({"error": pw_err}), 400
    if password != confirm_pw:
        return jsonify({"error": "Passwords do not match"}), 400
    if not _validate_drug_license(drug_license):
        return jsonify({"error": "Please enter a valid Drug License number (min 5 characters)"}), 400
    if not _validate_gstin(gstin):
        return jsonify({"error": "Invalid GSTIN format. Expected: 22AAAAA0000A1Z5"}), 400

    conn = get_db()
    if conn.execute("SELECT id FROM users WHERE email=%s", (email,)).fetchone():
        conn.close()
        return jsonify({"error": "An account with this email already exists"}), 409
    if conn.execute("SELECT id FROM users WHERE drug_license=%s", (drug_license,)).fetchone():
        conn.close()
        return jsonify({"error": "An account with this Drug License already exists"}), 409
    if conn.execute("SELECT id FROM users WHERE gstin=%s", (gstin,)).fetchone():
        conn.close()
        return jsonify({"error": "An account with this GSTIN already exists"}), 409

    user_id = uid()
    pw_hash = generate_password_hash(password)
    conn.execute("""
        INSERT INTO users (id, full_name, email, phone, pharmacy_type, password_hash, drug_license, gstin)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (user_id, full_name, email, phone, pharmacy_type, pw_hash, drug_license, gstin))

    # Create per-user settings row — store all registration fields immediately
    store_name = (wholesaler if is_wholesale else shop_name) or 'My Pharmacy'
    conn.execute("""
        INSERT INTO settings
          (store_name, store_type, address, phone, email, license_no, gstin,
           default_gst, low_stock_threshold, expiry_alert_days,
           owner_name, wholesaler, wholesaler_id,
           shop_name, retailer_owner, user_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (store_name, pharmacy_type, address, phone, email, drug_license, gstin,
          default_gst, low_stock, expiry_days,
          owner_name, wholesaler, wholesaler_id,
          shop_name, retailer_owner, user_id))
    conn.commit()

    # Give the new user their own copy of all seed medicines
    try:
        _copy_seed_products_to_user(conn, user_id)
    except Exception as e:
        print(f"  ⚠ Could not copy seed products to new user: {e}")

    conn.close()

    try:
        _send_welcome_email(email, full_name, pharmacy_type)
    except Exception:
        pass

    token = create_access_token(identity={
        'user_id': user_id, 'email': email, 'full_name': full_name,
        'pharmacy_type': pharmacy_type, 'drug_license': drug_license, 'gstin': gstin,
    })
    return jsonify({
        "ok": True,
        "message": "Account created successfully! Welcome to PharmaCare Pro.",
        "token": token,
        "user": {
            "id": user_id, "fullName": full_name, "email": email,
            "phone": phone, "pharmacyType": pharmacy_type,
            "drugLicense": drug_license, "gstin": gstin
        }
    }), 201


@app.route('/api/auth/login', methods=['POST'])
@require_json
def auth_login():
    d          = request.get_json()
    identifier = (d.get('identifier') or d.get('email') or '').strip()
    mode       = (d.get('mode') or 'auto').strip()   # 'gstin' | 'drug' | 'auto'
    password   = d.get('password') or ''

    if not identifier or not password:
        return jsonify({"error": "Please enter your credentials"}), 400

    # Validate GSTIN format when mode=gstin
    if mode == 'gstin' and not _validate_gstin(identifier.upper()):
        return jsonify({"error": "Invalid GSTIN format. Expected 15-character format e.g. 27ABCDE1234F1Z5"}), 400
    # Validate Drug License when mode=drug
    if mode == 'drug' and len(identifier.strip()) < 5:
        return jsonify({"error": "Drug License No. must be at least 5 characters"}), 400

    conn = get_db()
    # Match based on mode
    if mode == 'gstin':
        user = conn.execute("SELECT * FROM users WHERE UPPER(gstin)=%s", (identifier.upper(),)).fetchone()
    elif mode == 'drug':
        user = conn.execute("SELECT * FROM users WHERE UPPER(drug_license)=%s", (identifier.upper(),)).fetchone()
    else:
        # Auto-detect: try email, gstin, drug_license
        user = conn.execute("""
            SELECT * FROM users
            WHERE LOWER(email)=LOWER(%s) OR UPPER(drug_license)=UPPER(%s) OR UPPER(gstin)=UPPER(%s)
        """, (identifier, identifier, identifier)).fetchone()
    conn.close()

    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({"error": "Invalid credentials. Please check your Drug License / GSTIN / Email and password."}), 401

    # Seed copy: run in background thread so it never blocks login response.
    # _copy_seed_products_to_user is safe to call concurrently — it skips
    # medicines the user already has (checks by name+partition).
    def _seed_bg(uid):
        try:
            c = get_db()
            _copy_seed_products_to_user(c, uid)
            c.close()
        except Exception as e:
            print(f"  ⚠ Seed copy bg failed: {e}")

    concurrent.futures.ThreadPoolExecutor(max_workers=1).submit(_seed_bg, user['id'])

    conn2 = get_db()
    s_row = conn2.execute("SELECT * FROM settings WHERE user_id=%s", (user['id'],)).fetchone()
    s = dict(s_row) if s_row else {}
    conn2.close()

    token = create_access_token(identity={
        'user_id': user['id'],
        'email': user['email'],
        'full_name': user['full_name'],
        'pharmacy_type': user['pharmacy_type'],
        'drug_license': user['drug_license'],
        'gstin': user['gstin'],
    })

    return jsonify({
        "ok": True,
        "token": token,
        "user": {
            "id":           user['id'],
            "fullName":     user['full_name'],
            "email":        user['email'],
            "phone":        user['phone'],
            "pharmacyType": user['pharmacy_type'],
            "drugLicense":  user['drug_license'],
            "gstin":        user['gstin'],
            # Settings auto-populate fields
            "storeName":        s.get("store_name", ""),
            "shopName":         s.get("shop_name", ""),
            "retailerOwner":    s.get("retailer_owner", ""),
            "supplierName":     s.get("supplier_name", ""),
            "wholesaler":       s.get("wholesaler", ""),
            "ownerName":        s.get("owner_name", ""),
            "wholesalerId":     s.get("wholesaler_id", ""),
            "address":          s.get("address", ""),
            "defaultGst":       s.get("default_gst", 12),
            "currency":         s.get("currency", "₹"),
            "lowStockThreshold":s.get("low_stock_threshold", 10),
            "expiryAlertDays":  s.get("expiry_alert_days", 90),
            "wholesaleUpiQr":   s.get("wholesale_upi_qr", ""),
            "retailUpiQr":      s.get("retail_upi_qr", ""),
        }
    })


@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def auth_me():
    identity = _get_identity()
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=%s", (identity['user_id'],)).fetchone()
    conn.close()
    if not user:
        return jsonify({"error": "User not found"}), 404
    return jsonify({
        "id": user['id'],
        "fullName": user['full_name'],
        "email": user['email'],
        "phone": user['phone'],
        "pharmacyType": user['pharmacy_type'],
        "drugLicense": user['drug_license'],
        "gstin": user['gstin'],
        "createdAt": user['created_at'],
    })


@app.route('/api/auth/update-profile', methods=['PUT'])
@jwt_required()
@require_json
def auth_update_profile():
    identity = _get_identity()
    d = request.get_json()
    full_name = (d.get('fullName') or '').strip()
    phone     = (d.get('phone') or '').strip()
    email     = (d.get('email') or '').strip().lower()

    if not full_name:
        return jsonify({"error": "Full name is required"}), 400
    if not email or '@' not in email:
        return jsonify({"error": "Valid email is required"}), 400

    conn = get_db()
    # Check email uniqueness (exclude current user)
    existing = conn.execute(
        "SELECT id FROM users WHERE LOWER(email)=%s AND id!=%s",
        (email, identity['user_id'])
    ).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": "Email already in use by another account"}), 409

    conn.execute(
        "UPDATE users SET full_name=%s, phone=%s, email=%s WHERE id=%s",
        (full_name, phone, email, identity['user_id'])
    )
    conn.commit()
    conn.close()

    # Re-issue token with updated info
    new_token = create_access_token(identity={
        **identity,
        'full_name': full_name,
        'email': email,
    })
    return jsonify({"ok": True, "token": new_token})


@app.route('/api/auth/change-password', methods=['POST'])
@jwt_required()
@require_json
def auth_change_password():
    identity   = _get_identity()
    d          = request.get_json()
    current_pw = d.get('currentPassword') or ''
    new_pw     = d.get('newPassword') or ''
    confirm_pw = d.get('confirmPassword') or ''

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=%s", (identity['user_id'],)).fetchone()
    if not user or not check_password_hash(user['password_hash'], current_pw):
        conn.close()
        return jsonify({"error": "Current password is incorrect"}), 401

    pw_err = _validate_password(new_pw)
    if pw_err:
        conn.close()
        return jsonify({"error": pw_err}), 400
    if new_pw != confirm_pw:
        conn.close()
        return jsonify({"error": "New passwords do not match"}), 400

    conn.execute(
        "UPDATE users SET password_hash=%s WHERE id=%s",
        (generate_password_hash(new_pw), identity['user_id'])
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "message": "Password updated successfully"})


# ─────────────────────────────────────────────────────────────
# FORGOT-PASSWORD OTP STORE  (in-memory, 10-min expiry)
# ─────────────────────────────────────────────────────────────
_FP_OTP_STORE = {}  # { key: { otp, user_id, pw_hash, expires_at } }


def _fp_key(identifier: str) -> str:
    return identifier.strip().upper()


@app.route('/api/auth/forgot-password/send-otp', methods=['POST'])
@require_json
def fp_send_otp():
    """Send OTP to the registered email for password reset. No JWT needed."""
    d            = request.get_json()
    identifier   = (d.get('identifier') or '').strip()
    new_password = d.get('newPassword') or ''

    if not identifier:
        return jsonify({"error": "Please enter your GSTIN or Drug License No."}), 400

    pw_err = _validate_password(new_password)
    if pw_err:
        return jsonify({"error": pw_err}), 400

    conn = get_db()
    user = conn.execute("""
        SELECT * FROM users
        WHERE UPPER(gstin)=UPPER(%s) OR UPPER(drug_license)=UPPER(%s)
    """, (identifier, identifier)).fetchone()
    conn.close()

    # Return same response to prevent user enumeration
    if not user:
        return jsonify({"ok": True, "message": "If that account exists, an OTP has been sent."}), 200

    to_email = user['email']
    user_id  = user['id']
    otp      = _gen_otp()
    pw_hash  = generate_password_hash(new_password)
    key      = _fp_key(identifier)

    _FP_OTP_STORE[key] = {
        'otp':        otp,
        'user_id':    user_id,
        'pw_hash':    pw_hash,
        'expires_at': time.time() + 600,
    }

    otp_html = (
        "<html><body style='font-family:sans-serif;max-width:500px;margin:0 auto;padding:20px'>"
        "<div style='background:linear-gradient(135deg,#dc2626,#1e40af);padding:24px;"
        "border-radius:12px 12px 0 0;text-align:center'>"
        "<h2 style='color:white;margin:0'>PharmaCare Pro — Password Reset</h2></div>"
        "<div style='background:#f8fafc;padding:28px;border-radius:0 0 12px 12px;border:1px solid #e2e8f0'>"
        f"<p style='color:#374151'>Hello <strong>{user['full_name']}</strong>,</p>"
        "<p style='color:#374151'>Your password-reset OTP is:</p>"
        "<div style='background:white;border:2px dashed #dc2626;border-radius:12px;"
        "padding:20px;text-align:center;margin:16px 0'>"
        f"<span style='font-size:36px;font-weight:800;letter-spacing:8px;color:#dc2626;"
        f"font-family:monospace'>{otp}</span></div>"
        "<p style='color:#64748b;font-size:13px'>Expires in <strong>10 minutes</strong>. "
        "If you did not request this, ignore this email.</p>"
        "</div></body></html>"
    )
    email_sent = _send_email(to_email, f"PharmaCare Pro - Password Reset OTP: {otp}", otp_html)
    if not email_sent:
        print(f"  [FP-OTP] Email not sent. OTP for {to_email} ({identifier}): {otp}  (10 min)")

    resp = {"ok": True, "message": "OTP sent to registered email."}
    if not email_sent:
        resp["devOtp"] = otp   # auto-fill on frontend in dev mode
    return jsonify(resp)


@app.route('/api/auth/forgot-password/reset', methods=['POST'])
@require_json
def fp_reset_password():
    """Verify OTP and update password in DB. No JWT needed."""
    d            = request.get_json()
    identifier   = (d.get('identifier') or '').strip()
    otp_in       = (d.get('otp') or '').strip()
    new_password = d.get('newPassword') or ''

    if not identifier or not otp_in or not new_password:
        return jsonify({"error": "Missing required fields."}), 400

    key    = _fp_key(identifier)
    stored = _FP_OTP_STORE.get(key)

    if not stored:
        return jsonify({"error": "No OTP found. Please click \'Send OTP\' first."}), 400
    if time.time() > stored['expires_at']:
        del _FP_OTP_STORE[key]
        return jsonify({"error": "OTP has expired. Please request a new one."}), 400
    if otp_in != stored['otp']:
        return jsonify({"error": "Incorrect OTP. Please try again."}), 400

    user_id = stored['user_id']
    pw_hash = stored['pw_hash']
    del _FP_OTP_STORE[key]

    conn = get_db()
    conn.execute("UPDATE users SET password_hash=%s WHERE id=%s", (pw_hash, user_id))
    conn.commit()
    conn.close()

    print(f"  [FP] Password reset successful for user_id={user_id}")
    return jsonify({"ok": True, "message": "Password reset successfully!"})


# ─────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────


@app.route('/api/auth/send-otp', methods=['POST'])
@jwt_required()
@require_json
def send_otp():
    """Send OTP to new email when user wants to change email in settings."""
    identity  = _get_identity()
    user_id   = identity['user_id']
    d         = request.get_json()
    new_email = (d.get('newEmail') or '').strip().lower()

    if not new_email or '@' not in new_email:
        return jsonify({"error": "Valid email address required"}), 400

    # Check not already used
    conn = get_db()
    existing = conn.execute(
        "SELECT id FROM users WHERE LOWER(email)=%s AND id!=%s",
        (new_email, user_id)
    ).fetchone()
    conn.close()
    if existing:
        return jsonify({"error": "This email is already registered to another account"}), 409

    otp = _gen_otp()
    _OTP_STORE[user_id] = {
        'otp': otp,
        'new_email': new_email,
        'expires_at': time.time() + 600   # 10 minutes
    }

    # Send OTP email — never fails the request even if email breaks
    otp_html = (
        "<html><body style='font-family:sans-serif;max-width:500px;margin:0 auto;padding:20px'>"
        "<div style='background:linear-gradient(135deg,#1e40af,#0891b2);padding:24px;"
        "border-radius:12px 12px 0 0;text-align:center'>"
        "<h2 style='color:white;margin:0'>PharmaCare Pro</h2>"
        "<p style='color:#bfdbfe;margin:4px 0 0'>Email Verification</p></div>"
        "<div style='background:#f8fafc;padding:28px;border-radius:0 0 12px 12px;border:1px solid #e2e8f0'>"
        "<p style='color:#374151'>Your OTP for email change verification is:</p>"
        "<div style='background:white;border:2px dashed #1e40af;border-radius:12px;"
        "padding:20px;text-align:center;margin:16px 0'>"
        f"<span style='font-size:36px;font-weight:800;letter-spacing:8px;"
        f"color:#1e40af;font-family:monospace'>{otp}</span></div>"
        "<p style='color:#64748b;font-size:13px'>Expires in <strong>10 minutes</strong>. Do not share.</p>"
        "</div></body></html>"
    )
    email_sent = _send_email(new_email, f"PharmaCare Pro - OTP: {otp}", otp_html)
    if not email_sent:
        print(f"  [OTP] Email not sent — OTP for {new_email}: {otp}  (10 min)")

    resp_body = {"ok": True, "message": f"OTP sent to {new_email}"}
    if not email_sent:
        resp_body["devOtp"] = otp   # frontend fills the box automatically in dev mode
    return jsonify(resp_body)


@app.route('/api/auth/verify-otp', methods=['POST'])
@jwt_required()
@require_json
def verify_otp():
    """Verify OTP and update email if correct."""
    identity = _get_identity()
    user_id  = identity['user_id']
    d        = request.get_json()
    otp_in   = (d.get('otp') or '').strip()

    stored = _OTP_STORE.get(user_id)
    if not stored:
        return jsonify({"error": "No OTP found. Please request a new OTP."}), 400
    if time.time() > stored['expires_at']:
        del _OTP_STORE[user_id]
        return jsonify({"error": "OTP has expired. Please request a new one."}), 400
    if otp_in != stored['otp']:
        return jsonify({"error": "Incorrect OTP. Please try again."}), 400

    new_email = stored['new_email']
    del _OTP_STORE[user_id]

    conn = get_db()
    conn.execute("UPDATE users SET email=%s WHERE id=%s", (new_email, user_id))
    # Also update settings email
    conn.execute("UPDATE settings SET email=%s WHERE user_id=%s", (new_email, user_id))
    conn.commit()
    conn.close()

    # Re-issue JWT with new email
    new_token = create_access_token(identity={**identity, 'email': new_email})
    return jsonify({"ok": True, "message": "Email updated successfully!", "token": new_token, "newEmail": new_email})


@app.route('/api/settings', methods=['GET'])
@jwt_required()
def get_settings():
    identity = _get_identity()
    user_id  = identity['user_id']
    conn = get_db()
    # Try per-user settings first
    s = conn.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    if not s:
        s = conn.execute("SELECT * FROM settings ORDER BY id LIMIT 1").fetchone()
    user = _get_user_by_id(conn, user_id)
    conn.close()
    return jsonify(_settings_out_with_user(dict(s) if s else {}, user))


@app.route('/api/settings', methods=['PUT'])
@jwt_required()
@require_json
def save_settings():
    identity = _get_identity()
    user_id  = identity['user_id']
    d    = request.get_json()
    conn = get_db()
    # Get or create per-user settings
    existing = conn.execute("SELECT id FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    if not existing:
        conn.execute("INSERT INTO settings (user_id, store_type) VALUES (%s,%s)",
                     (user_id, identity.get('pharmacy_type', 'Retail Pharmacy')))

    conn.execute("""
        UPDATE settings SET
          store_name=%s, store_type=%s, address=%s, phone=%s, email=%s,
          license_no=%s, gstin=%s, default_gst=%s, currency=%s,
          low_stock_threshold=%s, expiry_alert_days=%s,
          wholesaler=%s, owner_name=%s, wholesaler_id=%s,
          shop_name=%s, retailer_owner=%s,
          wholesale_upi_qr=%s, retail_upi_qr=%s
        WHERE user_id=%s
    """, (
        d.get("storeName",         "My Pharmacy"),
        identity.get("pharmacy_type", "Retail Pharmacy"),  # LOCKED — always from JWT
        d.get("address",           ""),
        d.get("phone",             ""),
        d.get("email",             ""),
        identity.get("drug_license", ""),   # LOCKED — always from JWT
        identity.get("gstin",       ""),    # LOCKED — always from JWT
        float(d.get("defaultGst",  12)),
        d.get("currency",          "₹"),
        int(d.get("lowStockThreshold", 10)),
        int(d.get("expiryAlertDays",   90)),
        d.get("wholesaler",        ""),
        d.get("ownerName",         ""),
        d.get("wholesalerId",      ""),
        d.get("shopName",          ""),
        d.get("retailerOwner",     ""),
        d.get("wholesaleUpiQr",    ""),
        d.get("retailUpiQr",       ""),
        user_id,
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CATEGORIES  (global — no partition filtering)
# ─────────────────────────────────────────────────────────────
@app.route('/api/categories', methods=['GET'])
@jwt_required()
def get_categories():
    conn = get_db()
    rs   = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return jsonify(rows(rs))


@app.route('/api/categories', methods=['POST'])
@jwt_required()
@require_json
def add_category():
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Category name is required"}), 400
    conn   = get_db()
    exists = conn.execute(
        "SELECT id FROM categories WHERE LOWER(name)=%s", (name.lower(),)
    ).fetchone()
    if exists:
        conn.close()
        return jsonify({"error": "Category already exists"}), 409
    new_id = d.get("id") or uid()
    conn.execute(
        'INSERT INTO categories(id,name,"desc") VALUES(%s,%s,%s) ON CONFLICT (id) DO NOTHING',
        (new_id, name, d.get("desc", ""))
    )
    conn.commit()
    r = conn.execute("SELECT * FROM categories WHERE id=%s", (new_id,)).fetchone()
    conn.close()
    return jsonify(row(r)), 201


@app.route('/api/categories/<cat_id>', methods=['PUT'])
@jwt_required()
@require_json
def update_category(cat_id):
    """Edit an existing category name / description."""
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Category name is required"}), 400
    conn = get_db()
    # Check duplicate name (excluding self)
    exists = conn.execute(
        "SELECT id FROM categories WHERE LOWER(name)=%s AND id!=%s",
        (name.lower(), cat_id)
    ).fetchone()
    if exists:
        conn.close()
        return jsonify({"error": "Another category with that name already exists"}), 409
    conn.execute(
        'UPDATE categories SET name=%s, "desc"=%s WHERE id=%s',
        (name, d.get("desc", ""), cat_id)
    )
    conn.commit()
    r = conn.execute("SELECT * FROM categories WHERE id=%s", (cat_id,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "Category not found"}), 404
    return jsonify({"id": r["id"], "name": r["name"], "desc": r["desc"]})


@app.route('/api/categories/<cat_id>', methods=['DELETE'])
@jwt_required()
def delete_category(cat_id):
    conn  = get_db()
    count = conn.execute(
        "SELECT COUNT(*) FROM products WHERE category=%s", (cat_id,)
    ).fetchone()[0]
    if count > 0:
        conn.close()
        return jsonify({"error": f"Cannot delete: {count} medicine(s) use this category"}), 409
    conn.execute("DELETE FROM categories WHERE id=%s", (cat_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# MEDICINE DB — PRODUCTS (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/products', methods=['GET'])
@jwt_required()
def get_products():
    """
    Returns medicines for the CURRENT partition only.
    partition='both'  → visible to all modes (seed / migrated data)
    partition='wholesale' → only in Wholesale Pharma mode
    partition='retail'    → only in retail modes
    """
    conn = get_db()
    part = _jwt_partition()
    identity = _get_identity()
    user_id  = identity['user_id']
    rs   = conn.execute(
        "SELECT * FROM products WHERE user_id=%s AND partition IN (%s, 'both') ORDER BY name",
        (user_id, part)
    ).fetchall()
    conn.close()
    return jsonify([_product_out(r) for r in rs])


@app.route('/api/products', methods=['POST'])
@jwt_required()
@require_json
def add_product():
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Medicine name is required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    # Use provided partition or derive from current settings
    part   = d.get("partition") or _jwt_partition()
    conn.execute("""
        INSERT INTO products
          (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,"desc",
           partition,user_id,pieces_per_strip,strips_per_box,purchase_unit,selling_price)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        new_id, name,
        d.get("category", ""),
        d.get("unit",     "Tablet"),
        float(d.get("purchase",  0)),
        float(d.get("sale",      0)),
        float(d.get("gst",      12)),
        int(d.get("stock",       0)),
        int(d.get("minStock",   10)),
        d.get("sku",    "").strip(),
        d.get("expiry", ""),
        d.get("brand",  "").strip(),
        d.get("hsn",    "").strip(),
        d.get("desc",   "").strip(),
        part, identity['user_id'],
        int(d.get("piecesPerStrip", 10)),
        int(d.get("stripsPerBox",   10)),
        d.get("purchaseUnit", "strip"),
        float(d.get("sellingPrice", 0)),
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=%s", (new_id,)).fetchone()
    conn.close()
    return jsonify(_product_out(r)), 201


@app.route('/api/products/<prod_id>', methods=['PUT'])
@jwt_required()
@require_json
def update_product(prod_id):
    d    = request.get_json()
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Medicine name is required"}), 400
    conn = get_db()
    conn.execute("""
        UPDATE products SET
          name=%s, category=%s, unit=%s, purchase=%s, sale=%s, gst=%s,
          stock=%s, min_stock=%s, sku=%s, expiry=%s, brand=%s, hsn=%s, "desc"=%s,
          pieces_per_strip=%s, strips_per_box=%s, purchase_unit=%s, selling_price=%s
        WHERE id=%s
    """, (
        name,
        d.get("category", ""),
        d.get("unit",     "Tablet"),
        float(d.get("purchase",  0)),
        float(d.get("sale",      0)),
        float(d.get("gst",      12)),
        int(d.get("stock",       0)),
        int(d.get("minStock",   10)),
        d.get("sku",    "").strip(),
        d.get("expiry", ""),
        d.get("brand",  "").strip(),
        d.get("hsn",    "").strip(),
        d.get("desc",   "").strip(),
        int(d.get("piecesPerStrip", 10)),
        int(d.get("stripsPerBox",   10)),
        d.get("purchaseUnit", "strip"),
        float(d.get("sellingPrice", 0)),
        prod_id,
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=%s", (prod_id,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "Product not found"}), 404
    return jsonify(_product_out(r))


@app.route('/api/products/<prod_id>', methods=['DELETE'])
@jwt_required()
def delete_product(prod_id):
    conn = get_db()
    conn.execute("DELETE FROM products WHERE id=%s", (prod_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route('/api/products/<prod_id>/stock', methods=['PATCH'])
@jwt_required()
@require_json
def adjust_stock(prod_id):
    """mode = 'add' | 'remove' | 'set',  qty = integer."""
    d    = request.get_json()
    mode = d.get("mode", "add")
    qty  = int(d.get("qty", 0))
    conn = get_db()
    p    = conn.execute("SELECT stock FROM products WHERE id=%s", (prod_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify({"error": "Product not found"}), 404
    current   = p["stock"]
    new_stock = (current + qty if mode == "add"
                 else max(0, current - qty) if mode == "remove"
                 else max(0, qty))
    conn.execute("UPDATE products SET stock=%s WHERE id=%s", (new_stock, prod_id))
    conn.commit()
    r = conn.execute("SELECT * FROM products WHERE id=%s", (prod_id,)).fetchone()
    conn.close()
    return jsonify(_product_out(r))


# ─────────────────────────────────────────────────────────────
# MEDICINE DB — STOCK-IN (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/stock-ins', methods=['GET'])
@jwt_required()
def get_stock_ins():
    identity = _get_identity()   # ADD THIS LINE
    conn = get_db()
    part = _jwt_partition()
    rs   = conn.execute(
        "SELECT * FROM stock_ins WHERE user_id=%s AND partition IN (%s, 'both') ORDER BY date DESC LIMIT 200",
        (identity['user_id'], part)
    ).fetchall()
    conn.close()
    return jsonify([{
        "id":          r["id"],
        "date":        r["date"],
        "productId":   r["product_id"],
        "productName": r["product_name"],
        "qty":         r["qty"],
        "price":       r["price"],
        "batch":       r["batch"],
        "expiry":      r["expiry"],
        "supplier":    r["supplier"],
        "invoiceNo":   r["invoice_no"],
        "notes":       r["notes"],
        "partition":   r["partition"] if "partition" in r.keys() else PARTITION_BOTH,
    } for r in rs])


@app.route('/api/stock-ins', methods=['POST'])
@jwt_required()
@require_json
def add_stock_in():
    d       = request.get_json()
    prod_id = d.get("productId", "")
    qty     = int(d.get("qty", 0))
    price   = float(d.get("price", 0))
    batch   = d.get("batch",  "").strip()
    expiry  = d.get("expiry", "")

    if not prod_id or qty < 1:
        return jsonify({"error": "Product and valid quantity are required"}), 400

    conn   = get_db()
    new_id = d.get("id") or uid()
    part   = d.get("partition") or _jwt_partition()

    # Replace the broken INSERT block:
    identity = _get_identity()
    conn.execute("""
    INSERT INTO stock_ins
      (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition,user_id)
    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
""", (
    new_id,
    d.get("date") or today_str(),
    prod_id,
    d.get("productName", ""),
    qty, price, batch, expiry,
    d.get("supplier",  "").strip(),
    d.get("invoiceNo", "").strip(),
    d.get("notes",     "").strip(),
    part, identity['user_id']
    ))

    
    conn.commit()
    updated = conn.execute("SELECT * FROM products WHERE id=%s", (prod_id,)).fetchone()
    conn.close()
    return jsonify({
        "ok": True,
        "updatedProduct": _product_out(updated) if updated else None,
    }), 201


# ─────────────────────────────────────────────────────────────
# BILLS — calculate (no DB write)
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills/calculate', methods=['POST'])
@jwt_required()
@require_json
def calculate_bill():
    d         = request.get_json()
    raw_items = d.get("items", [])
    out_items = []
    for it in raw_items:
        qty        = float(it.get("qty",       1))
        unit_price = float(it.get("unitPrice", 0))
        discount   = float(it.get("discount",  0))
        gst_rate   = float(it.get("gstRate",   0))
        gst_amt, line_total = _calc_item(unit_price, qty, discount, gst_rate)
        out_items.append({
            **it,
            "qty":       qty,
            "unitPrice": unit_price,
            "discount":  discount,
            "gstRate":   gst_rate,
            "gstAmt":    gst_amt,
            "lineTotal": line_total,
        })
    totals = _calc_totals([{
        "qty":        it["qty"],
        "unit_price": it["unitPrice"],
        "discount":   it["discount"],
        "gst_amt":    it["gstAmt"],
    } for it in out_items])
    return jsonify({"items": out_items, "totals": totals})


# ─────────────────────────────────────────────────────────────
# BILLS — next number
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills/next-number', methods=['GET'])
@jwt_required()
def next_bill_number():
    identity = _get_identity()
    user_id  = identity['user_id']
    conn     = get_db()
    s_row    = conn.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    if not s_row:
        s_row = conn.execute("SELECT * FROM settings ORDER BY id LIMIT 1").fetchone()
    s   = dict(s_row) if s_row else {}
    n   = s.get("next_bill_no", 1)
    fmt = str(n).zfill(4)
    conn.close()
    return jsonify({"nextBillNo": n, "formatted": fmt, "gstin": identity.get("gstin", "")})


# ─────────────────────────────────────────────────────────────
# SALES HISTORY DB — BILLS list & get (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills', methods=['GET'])
@jwt_required()
def get_bills():
    identity = _get_identity()        # ADD THIS
    user_id  = identity['user_id']    # ADD THIS
    conn   = get_db()
    part   = _jwt_partition()
    bill_w = _bill_type_filter(part)

    query  = f"SELECT * FROM bills WHERE user_id=%s AND {bill_w}"
    params = [user_id]
    

    q     = request.args.get("q",       "")
    frm   = request.args.get("from",    "")
    to    = request.args.get("to",      "")
    pay   = request.args.get("payment", "")
    btype = request.args.get("type",    "")
    limit = int(request.args.get("limit", 500))

    if q:
        query  += " AND (bill_no LIKE %s OR customer LIKE %s OR doctor LIKE %s OR phone LIKE %s)"
        p       = f"%{q}%"
        params += [p, p, p, p]
    if frm:
        query  += " AND date >= %s"; params.append(frm)
    if to:
        query  += " AND date <= %s"; params.append(to)
    if pay:
        query  += " AND payment_mode = %s"; params.append(pay)
    if btype:
        query  += " AND bill_store_type = %s"; params.append(btype)

    query  += " ORDER BY created_at DESC LIMIT %s"
    params.append(limit)

    bill_rows = conn.execute(query, params).fetchall()
    result    = [_bill_out(b, conn) for b in bill_rows]
    conn.close()
    return jsonify(result)


@app.route('/api/bills/<bill_id>', methods=['GET'])
@jwt_required()
def get_bill(bill_id):
    conn = get_db()
    b    = conn.execute("SELECT * FROM bills WHERE id=%s", (bill_id,)).fetchone()
    if not b:
        conn.close()
        return jsonify({"error": "Bill not found"}), 404
    result = _bill_out(b, conn)
    conn.close()
    return jsonify(result)


# ─────────────────────────────────────────────────────────────
# SALES HISTORY DB — BILLS create
# ─────────────────────────────────────────────────────────────
@app.route('/api/bills', methods=['POST'])
@jwt_required()
@require_json
def create_bill():
    identity  = _get_identity()
    user_id   = identity['user_id']
    d         = request.get_json()
    conn      = get_db()
    s_row = conn.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    s = dict(s_row) if s_row else _get_settings(conn)
    raw_items = d.get("items", [])

    if not raw_items:
        conn.close()
        return jsonify({"error": "Bill must have at least one item"}), 400

    calc_items = []
    # Build product lookup for unit/pack data + purchase price
    all_prod_ids = [it.get("productId", "") for it in raw_items if it.get("productId")]
    prod_data_map = {}
    if all_prod_ids:
        ph2 = ",".join(["%s"] * len(all_prod_ids))
        for row in conn.execute(
            f"SELECT id, purchase, selling_price, pieces_per_strip, strips_per_box, purchase_unit FROM products WHERE id IN ({ph2})",
            all_prod_ids
        ).fetchall():
            prod_data_map[row["id"]] = {
                "purchase":       float(row["purchase"] or 0),
                "sellingPrice":   float(row["selling_price"] or 0),
                "piecesPerStrip": int(row["pieces_per_strip"] or 10),
                "stripsPerBox":   int(row["strips_per_box"]   or 10),
                "purchaseUnit":   row["purchase_unit"] or "strip",
            }

    for it in raw_items:
        qty        = float(it.get("qty",       1))
        unit_price = float(it.get("unitPrice", 0))
        discount   = float(it.get("discount",  0))
        gst_rate   = float(it.get("gstRate",   0))
        unit_type  = (it.get("unitType") or "strip").lower()   # box | strip | piece
        display_qty = qty  # what the user typed (e.g. "2" strips)

        # Normalise to pieces based on unit_type
        pd = prod_data_map.get(it.get("productId", ""), {})
        pps = pd.get("piecesPerStrip", 10)  # pieces per strip
        spb = pd.get("stripsPerBox",   10)  # strips per box
        if unit_type == "box":
            qty_in_pieces = display_qty * spb * pps
        elif unit_type == "strip":
            qty_in_pieces = display_qty * pps
        else:  # piece
            qty_in_pieces = display_qty

        gst_amt, line_total = _calc_item(unit_price, qty, discount, gst_rate)
        calc_items.append({
            "productId":   it.get("productId", ""),
            "name":        it.get("name",      ""),
            "category":    it.get("category",  ""),
            "unit":        it.get("unit",       ""),
            "qty":         qty,           # display qty (in selected unit)
            "unitPrice":   unit_price,
            "discount":    discount,
            "gstRate":     gst_rate,
            "gstAmt":      gst_amt,
            "lineTotal":   line_total,
            "unitType":    unit_type,
            "displayQty":  display_qty,
            "qtyInPieces": qty_in_pieces,
        })

    totals = _calc_totals([{
        "qty":        it["qty"],
        "unit_price": it["unitPrice"],
        "discount":   it["discount"],
        "gst_amt":    it["gstAmt"],
    } for it in calc_items])

    n   = s.get("next_bill_no", 1)
    fmt = str(n).zfill(4)
    is_ws    = d.get("billStoreType", "retail") == "wholesale"
    ws_gstin = d.get("wsGstin", "") or identity.get("gstin", "")
    bill_no  = f"{ws_gstin}-{fmt}" if is_ws and ws_gstin else fmt

    bill_id = d.get("id") or uid()
    conn.execute("""
        INSERT INTO bills
          (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
           subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
           ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
           rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone,user_id)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        bill_id, bill_no,
        d.get("date") or today_str(),
        d.get("customer",        "Walk-in"),
        d.get("phone",           ""),
        d.get("doctor",          ""),
        d.get("rx",              ""),
        d.get("paymentMode",     "Cash"),
        d.get("notes",           ""),
        totals["subtotal"],
        totals["total_discount"],
        totals["total_gst"],
        totals["round_off"],
        totals["grand_total"],
        "wholesale" if is_ws else "retail",
        d.get("wsSupplier",      ""),
        d.get("wsOwner",         ""),
        ws_gstin,
        d.get("shopName",        ""),
        d.get("shopkeeperGstin", ""),
        d.get("rtShop",          ""),
        d.get("rtOwner",         ""),
        d.get("rtGstin",         ""),
        d.get("rtLicense",       ""),
        d.get("rtEmail",         ""),
        d.get("rtPhone",         ""),
        user_id,                          # ← was missing: bills had user_id=NULL
    ))

    # prod_data_map was already built during calc_items — reuse it for purchase price
    low_stock_alerts = []
    for it in calc_items:
        pd_item = prod_data_map.get(it["productId"], {})
        purchase_price_raw = float(pd_item.get("purchase", 0.0) or 0.0)
        purchase_unit      = (pd_item.get("purchaseUnit", "strip") or "strip").lower()
        pps                = int(pd_item.get("piecesPerStrip", 10) or 10) or 1
        spb                = int(pd_item.get("stripsPerBox",   10) or 10) or 1
        unit_type          = (it.get("unitType", "strip") or "strip").lower()

        # ── Normalise purchase_price to the same unit as unit_price (billing unit) ──
        # Step 1: convert raw purchase price → cost per PIECE
        if purchase_unit == "box":
            cost_per_piece = purchase_price_raw / (spb * pps)
        elif purchase_unit == "strip":
            cost_per_piece = purchase_price_raw / pps
        else:                                   # "piece" — already per piece
            cost_per_piece = purchase_price_raw

        # Step 2: convert cost-per-piece → cost per BILLING UNIT
        if unit_type == "box":
            purchase_price = cost_per_piece * spb * pps
        elif unit_type == "strip":
            purchase_price = cost_per_piece * pps
        else:                                   # "piece"
            purchase_price = cost_per_piece
        # ────────────────────────────────────────────────────────────────────────────
        # Now purchase_price and unit_price are both expressed per billing unit,
        # so  profit = (unit_price × (1 – disc%) – purchase_price) × qty  is correct.

        selling_price_snap = pd_item.get("sellingPrice", 0.0)
        conn.execute("""
            INSERT INTO bill_items
              (id,bill_id,product_id,name,category,unit,qty,unit_price,
               discount,gst_rate,gst_amt,line_total,purchase_price,
               unit_type,display_qty,qty_in_pieces)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            uid(), bill_id,
            it["productId"], it["name"], it["category"], it["unit"],
            it["qty"], it["unitPrice"], it["discount"],
            it["gstRate"], it["gstAmt"], it["lineTotal"],
            purchase_price,
            it["unitType"], it["displayQty"], it["qtyInPieces"],
        ))
        if it["productId"]:
            # Deduct in PIECES — accurate regardless of sale unit (box/strip/piece)
            conn.execute(
                "UPDATE products SET stock = GREATEST(0, stock - %s) WHERE id=%s",
                (int(it["qtyInPieces"]), it["productId"])
            )
            p = conn.execute(
                "SELECT name, stock, min_stock FROM products WHERE id=%s",
                (it["productId"],)
            ).fetchone()
            if p and p["stock"] <= p["min_stock"]:
                low_stock_alerts.append({"name": p["name"], "stock": p["stock"]})

    conn.execute("UPDATE settings SET next_bill_no = next_bill_no + 1 WHERE user_id=%s", (user_id,))
    conn.commit()

    saved  = conn.execute("SELECT * FROM bills WHERE id=%s", (bill_id,)).fetchone()
    result = _bill_out(saved, conn)
    new_n  = n + 1
    conn.close()

    return jsonify({
        "bill":           result,
        "nextBillNo":     new_n,
        "lowStockAlerts": low_stock_alerts,
    }), 201


@app.route('/api/bills/<bill_id>', methods=['DELETE'])
@jwt_required()
def delete_bill(bill_id):
    conn = get_db()
    conn.execute("DELETE FROM bill_items WHERE bill_id=%s", (bill_id,))
    conn.execute("DELETE FROM bills WHERE id=%s",           (bill_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CREDIT DB — WHOLESALE credits (partition='wholesale'|'both')
# ─────────────────────────────────────────────────────────────
@app.route('/api/credits', methods=['GET'])
@jwt_required()
def get_credits():
    """
    Returns credits for Wholesale Pharma partition only.
    These represent retailers/shops that owe money to this wholesaler.
    """
    conn = get_db()
    rs   = conn.execute(
        "SELECT * FROM credits WHERE user_id=%s AND partition IN ('wholesale', 'both') ORDER BY date DESC",
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify([_credit_out(r) for r in rs])


@app.route('/api/credits', methods=['POST'])
@jwt_required()
@require_json
def add_credit():
    d = request.get_json()
    if not d.get("shopName") or not d.get("shopkeeperName"):
        return jsonify({"error": "Shop name and shopkeeper name are required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    conn.execute("""
        INSERT INTO credits
          (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        new_id,
        d.get("date") or today_str(),
        d.get("shopName",       ""),
        d.get("shopkeeperName", ""),
        d.get("phone",          ""),
        d.get("forItem",        ""),
        float(d.get("amount",   0)),
        d.get("method",         "Cash"),
        d.get("status",         "Pending"),
        PARTITION_WS,   # always wholesale
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM credits WHERE id=%s", (new_id,)).fetchone()
    conn.close()
    return jsonify(_credit_out(r)), 201


@app.route('/api/credits/bulk', methods=['DELETE'])
@jwt_required()
def bulk_delete_credits():
    """
    Delete credits for the authenticated user within a date range.
    Query param: period = '7' | '30' | '90'
    '7'  → last 7 calendar days
    '30' → current calendar month to date
    '90' → last 90 calendar days
    """
    identity = _get_identity()
    user_id  = identity['user_id']
    period   = request.args.get('period', '')
    from datetime import date as _date, timedelta as _td

    today = _date.today()
    if period == '7':
        cutoff = today - _td(days=7)
    elif period == '30':
        cutoff = _date(today.year, today.month, 1)
    elif period == '90':
        cutoff = today - _td(days=90)
    else:
        return jsonify({"error": "period must be 7, 30, or 90"}), 400

    cutoff_str = cutoff.isoformat()   # 'YYYY-MM-DD'
    conn = get_db()
    result = conn.execute(
        "DELETE FROM credits WHERE user_id=%s AND date >= %s",
        (user_id, cutoff_str)
    )
    deleted = result.rowcount if result else 0
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "deleted": deleted})


@app.route('/api/credits/<credit_id>', methods=['PATCH'])
@jwt_required()
@require_json
def update_credit_status(credit_id):
    d    = request.get_json()
    conn = get_db()
    conn.execute(
        "UPDATE credits SET status=%s WHERE id=%s",
        (d.get("status", "Pending"), credit_id)
    )
    conn.commit()
    r = conn.execute("SELECT * FROM credits WHERE id=%s", (credit_id,)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "Credit not found"}), 404
    return jsonify(_credit_out(r))


@app.route('/api/credits/<credit_id>', methods=['DELETE'])
@jwt_required()
def delete_credit(credit_id):
    conn = get_db()
    conn.execute("DELETE FROM credits WHERE id=%s", (credit_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CREDIT DB — RETAIL shop_credits (partition='retail'|'both')
# ─────────────────────────────────────────────────────────────
@app.route('/api/shop-credits', methods=['GET'])
@jwt_required()
def get_shop_credits():
    """
    Returns shop credits for retail partition only.
    These represent what this retail shop owes to suppliers/wholesalers.
    """
    conn = get_db()
    rs   = conn.execute(
        "SELECT * FROM shop_credits WHERE user_id=%s AND partition IN ('retail', 'both') ORDER BY bill_date DESC",
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify([_shop_credit_out(r) for r in rs])


@app.route('/api/shop-credits/fetch/<supplier_id>', methods=['GET'])
@jwt_required()
def fetch_shop_credit_by_supplier(supplier_id):
    conn = get_db()
    r    = conn.execute("""
        SELECT * FROM shop_credits
        WHERE LOWER(supplier_id)=%s AND partition IN ('retail', 'both')
        ORDER BY bill_date DESC
        LIMIT 1
    """, (supplier_id.lower(),)).fetchone()
    conn.close()
    if not r:
        return jsonify({"error": "No record found"}), 404
    return jsonify(_shop_credit_out(r))


@app.route('/api/shop-credits', methods=['POST'])
@jwt_required()
@require_json
def add_shop_credit():
    d = request.get_json()
    if not d.get("supplierId") or not d.get("supplierName"):
        return jsonify({"error": "Supplier ID and name are required"}), 400
    new_id = d.get("id") or uid()
    conn   = get_db()
    conn.execute("""
        INSERT INTO shop_credits
          (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
           payment_mode,pending,last_purchase_date,bill_date,status,partition)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        new_id,
        d.get("supplierId",       ""),
        d.get("supplierName",     ""),
        d.get("ownerName",        ""),
        float(d.get("totalPurchase", 0)),
        float(d.get("paid",          0)),
        d.get("paymentMode",      "Cash"),
        float(d.get("pending",       0)),
        d.get("lastPurchaseDate") or today_str(),
        d.get("billDate")         or today_str(),
        d.get("status",           "Pending"),
        PARTITION_RT,   # always retail
    ))
    conn.commit()
    r = conn.execute("SELECT * FROM shop_credits WHERE id=%s", (new_id,)).fetchone()
    conn.close()
    return jsonify(_shop_credit_out(r)), 201


@app.route('/api/shop-credits/<sc_id>', methods=['DELETE'])
@jwt_required()
def delete_shop_credit(sc_id):
    conn = get_db()
    conn.execute("DELETE FROM shop_credits WHERE id=%s", (sc_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route('/api/shop-credits/supplier/<supplier_id>/history', methods=['GET'])
@jwt_required()
def supplier_history(supplier_id):
    conn = get_db()
    rs   = conn.execute("""
        SELECT * FROM shop_credits
        WHERE LOWER(supplier_id)=%s AND partition IN ('retail', 'both')
        ORDER BY bill_date DESC
    """, (supplier_id.lower(),)).fetchall()
    conn.close()
    if not rs:
        return jsonify({"error": "No records found"}), 404
    records         = [_shop_credit_out(r) for r in rs]
    total_purchased = round(sum(r["totalPurchase"] for r in records), 2)
    total_paid      = round(sum(r["paid"]          for r in records), 2)
    current_pending = records[0]["pending"] if records else 0
    return jsonify({
        "records":        records,
        "totalPurchased": total_purchased,
        "totalPaid":      total_paid,
        "currentPending": current_pending,
        "supplierName":   records[0]["supplierName"] if records else "",
        "ownerName":      records[0]["ownerName"]    if records else "",
    })


# ─────────────────────────────────────────────────────────────
# DASHBOARD (partition-aware)
# ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    identity   = _get_identity()
    user_id    = identity['user_id']
    conn       = get_db()
    # Load per-user settings
    s_raw = conn.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    s     = dict(s_raw) if s_raw else {}
    today      = today_str()
    alert_days = s.get("expiry_alert_days", 90)
    store_type = (identity.get('pharmacy_type') or s.get("store_type", "Retail Pharmacy")).strip()
    part       = _store_partition(store_type)
    bill_where = _bill_type_filter(part)
    bill_type_key = part  # 'wholesale' | 'retail'

    # Get reset date for this user + partition
    reset_row = conn.execute(
        "SELECT reset_date FROM dashboard_resets WHERE store_type_key=%s AND user_id=%s",
        (bill_type_key, user_id)
    ).fetchone()
    reset_date = reset_row["reset_date"] if reset_row else None

    date_filter = ""
    date_params_today = [today]
    if reset_date:
        date_filter = " AND date >= %s"
        date_params_today.append(reset_date)

    # Today revenue (partition-filtered)
    today_bills = conn.execute(
        f"SELECT grand_total FROM bills WHERE user_id=%s AND date=%s AND {bill_where}{date_filter}",
        [user_id] + date_params_today
    ).fetchall()
    today_rev = sum(r["grand_total"] for r in today_bills)

    # Products — partition-filtered for stock/expiry alerts
    products      = conn.execute(
        "SELECT * FROM products WHERE user_id=%s AND partition IN (%s, 'both')", (user_id, part)
    ).fetchall()
    low_stock     = [_product_out(p) for p in products if p["stock"] <= p["min_stock"]]
    expiry_alerts, expired_count = [], 0
    for p in products:
        days = expiry_days_left(p["expiry"])
        if days < 0:
            expired_count += 1
            expiry_alerts.append({**_product_out(p), "daysLeft": days})
        elif days <= alert_days:
            expiry_alerts.append({**_product_out(p), "daysLeft": days})

    # Revenue last 7 days
    rev_7 = []
    for i in range(6, -1, -1):
        d     = (date.today() - timedelta(days=i)).isoformat()
        params = [d]
        extra  = ""
        if reset_date and d < reset_date:
            rev_7.append({"date": d, "revenue": 0.0})
            continue
        if reset_date:
            extra = " AND date >= %s"
            params.append(reset_date)
        total = float(conn.execute(
            f"SELECT COALESCE(SUM(grand_total),0) FROM bills WHERE user_id=%s AND date=%s AND {bill_where}{extra}",
            [user_id] + params
        ).fetchone()[0])
        rev_7.append({"date": d, "revenue": round(float(total), 2)})

    # Top 8 products by pieces sold (qty_in_pieces normalises box/strip/piece to pieces)
    reset_clause = f" AND b.date >= '{reset_date}'" if reset_date else ""
    top_rows = conn.execute(f"""
        SELECT bi.name, SUM(COALESCE(bi.qty_in_pieces, bi.qty)) as units
        FROM bill_items bi
        JOIN bills b ON b.id = bi.bill_id
        WHERE b.user_id=%s AND b.{bill_where}{reset_clause}
        GROUP BY bi.name ORDER BY units DESC LIMIT 8
    """, (user_id,)).fetchall()
    top_products = [{"name": r["name"], "units": r["units"]} for r in top_rows]

    # Weekly profit (current month, partition-filtered)
    now         = date.today()
    month_start = now.replace(day=1).isoformat()
    next_mo     = (now.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end   = (next_mo - timedelta(days=1)).isoformat()
    effective_start = max(month_start, reset_date) if reset_date else month_start
    bills_month  = conn.execute(
        f"SELECT id, date FROM bills WHERE user_id=%s AND date>=%s AND date<=%s AND {bill_where}",
        (user_id, effective_start, month_end)
    ).fetchall()
    week_profit = [0.0, 0.0, 0.0, 0.0]
    for b in bills_month:
        day_n = int(b["date"].split("-")[2])
        w_idx = 0 if day_n <= 7 else 1 if day_n <= 14 else 2 if day_n <= 21 else 3
        for it in conn.execute(
            """SELECT bi.unit_price, bi.qty, bi.discount, bi.purchase_price,
                      COALESCE(bi.unit_type,'strip')   AS unit_type,
                      COALESCE(bi.qty_in_pieces, bi.qty) AS qty_in_pieces,
                      p.purchase       AS prod_purchase,
                      COALESCE(p.purchase_unit,'strip') AS purchase_unit,
                      COALESCE(p.pieces_per_strip,10)   AS pps,
                      COALESCE(p.strips_per_box,10)     AS spb
               FROM bill_items bi
               LEFT JOIN products p ON bi.product_id = p.id
               WHERE bi.bill_id=%s""",
            (b["id"],)
        ).fetchall():
            qty        = float(it["qty"] or 1)
            unit_price = float(it["unit_price"] or 0)
            discount   = float(it["discount"]   or 0)
            sale_disc  = unit_price * (1 - discount / 100)

            # Pieces per billing unit  (stored at insert time — always accurate)
            qty_pieces = float(it["qty_in_pieces"] or qty)
            ppu        = (qty_pieces / qty) if qty > 0 else 1   # pieces per billing unit

            prod_purchase = float(it["prod_purchase"] or 0)
            pps = int(it["pps"] or 10) or 1
            spb = int(it["spb"] or 10) or 1
            purchase_unit = (it["purchase_unit"] or "strip").lower()

            if prod_purchase > 0:
                # Re-derive cost-per-piece from the product's current purchase price.
                # This corrects legacy rows where purchase_price was stored as per-box.
                if purchase_unit == "box":
                    cost_per_piece = prod_purchase / (spb * pps)
                elif purchase_unit == "strip":
                    cost_per_piece = prod_purchase / pps
                else:
                    cost_per_piece = prod_purchase
                purchase = cost_per_piece * ppu          # cost in billing unit
            else:
                # Product deleted — fall back to stored value.
                # New bills store this correctly (per billing unit); old bills may be
                # per-box, but there is no product row left to re-derive from.
                purchase = float(it["purchase_price"] or 0)

            week_profit[w_idx] += (sale_disc - purchase) * qty
    week_profit = [round(v, 2) for v in week_profit]

    # Recent 8 bills
    reset_q = f" AND date >= '{reset_date}'" if reset_date else ""
    recent_rows  = conn.execute(
        f"SELECT * FROM bills WHERE user_id=%s AND {bill_where}{reset_q} ORDER BY created_at DESC LIMIT 8",
        (user_id,)
    ).fetchall()
    recent_bills = [_bill_out(b, conn) for b in recent_rows]

    conn.close()
    return jsonify({
        "todayRevenue":   round(today_rev, 2),
        "todayBillCount": len(today_bills),
        "totalProducts":  len(products),
        "lowStockCount":  len(low_stock),
        "lowStockItems":  low_stock[:10],
        "expiryAlerts":   expiry_alerts[:6],
        "expiredCount":   expired_count,
        "revenue7Days":   rev_7,
        "topProducts":    top_products,
        "weekProfit":     week_profit,
        "recentBills":    recent_bills,
        "products":       [_product_out(p) for p in products],
        "resetDate":      reset_date,
        "storeTypeKey":   bill_type_key,
    })


# ─────────────────────────────────────────────────────────────
# DASHBOARD RESETS
# ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard/resets', methods=['GET'])
@jwt_required()
def get_dashboard_resets():
    conn = get_db()
    identity = _get_identity()
    rs   = conn.execute("SELECT store_type_key, reset_date FROM dashboard_resets WHERE user_id=%s", (identity['user_id'],)).fetchall()
    conn.close()
    return jsonify({"resets": [{"storeTypeKey": r["store_type_key"], "resetDate": r["reset_date"]} for r in rs]})


@app.route('/api/dashboard/reset', methods=['POST'])
@jwt_required()
def post_dashboard_reset():
    d     = request.get_json(force=True) or {}
    key   = d.get("storeTypeKey", "").strip()
    rdate = d.get("resetDate", "").strip()
    if key not in ("wholesale", "retail"):
        return jsonify({"error": "Invalid storeTypeKey"}), 400
    if not rdate:
        return jsonify({"error": "resetDate required"}), 400
    conn = get_db()
    identity = _get_identity()
    user_id  = identity['user_id']
    conn.execute(
        "INSERT INTO dashboard_resets (store_type_key, reset_date, user_id) VALUES (%s,%s,%s) "
        "ON CONFLICT (store_type_key, user_id) DO UPDATE SET reset_date=EXCLUDED.reset_date",
        (key, rdate, user_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "storeTypeKey": key, "resetDate": rdate})


# ─────────────────────────────────────────────────────────────
# ANALYSIS (partition-aware)
# ─────────────────────────────────────────────────────────────
@app.route('/api/analysis', methods=['GET'])
@jwt_required()
def get_analysis():
    days      = int(request.args.get("days", 7))
    from_date = (date.today() - timedelta(days=days)).isoformat()
    conn      = get_db()
    part      = _jwt_partition()
    bill_w    = _bill_type_filter(part)

    identity  = _get_identity()
    user_id   = identity['user_id']
    bills     = conn.execute(
        f"SELECT * FROM bills WHERE user_id=%s AND date>=%s AND {bill_w}", (user_id, from_date)
    ).fetchall()
    bill_ids  = [b["id"] for b in bills]

    total_rev  = sum(b["grand_total"] for b in bills)
    avg_bill   = (total_rev / len(bills)) if bills else 0
    prod_sales, cat_sales, pay_totals, rev_by_day = {}, {}, {}, {}

    for b in bills:
        pm = b["payment_mode"]
        pay_totals[pm]    = round(pay_totals.get(pm, 0) + b["grand_total"], 2)
        rev_by_day[b["date"]] = round(rev_by_day.get(b["date"], 0) + b["grand_total"], 2)

    if bill_ids:
        ph    = ",".join(["%s"] * len(bill_ids))
        items = conn.execute(
            f"SELECT * FROM bill_items WHERE bill_id IN ({ph})", bill_ids
        ).fetchall()
        cats  = {r["id"]: r["name"] for r in
                 conn.execute("SELECT id, name FROM categories WHERE user_id=%s", (user_id,)).fetchall()}
        for it in items:
            n  = it["name"]
            cn = cats.get(it["category"], "Uncategorized")
            if n not in prod_sales:
                prod_sales[n] = {"units": 0, "revenue": 0, "category": it["category"]}
            # Use qty_in_pieces (pieces-normalised) so strip/box/piece sales compare fairly
            prod_sales[n]["units"]   += float(it["qty_in_pieces"] or it["qty"] or 0)
            prod_sales[n]["revenue"]  = round(prod_sales[n]["revenue"] + it["line_total"], 2)
            cat_sales[cn] = round(cat_sales.get(cn, 0) + it["line_total"], 2)

    sorted_prods = sorted(prod_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)
    top_product  = sorted_prods[0][0] if sorted_prods else "—"

    rev_list = []
    for i in range(days - 1, -1, -1):
        d_str = (date.today() - timedelta(days=i)).isoformat()
        rev_list.append({"date": d_str, "revenue": rev_by_day.get(d_str, 0)})

    conn.close()
    return jsonify({
        "totalBills":       len(bills),
        "totalRevenue":     round(total_rev, 2),
        "avgBillValue":     round(avg_bill, 2),
        "topProduct":       top_product,
        "productSales":     [{"name": k, **v} for k, v in sorted_prods],
        "categorySales":    [{"name": k, "revenue": v} for k, v in
                             sorted(cat_sales.items(), key=lambda x: -x[1])],
        "paymentBreakdown": [{"mode": k, "total": v} for k, v in pay_totals.items()],
        "revenueByDay":     rev_list,
    })


# ─────────────────────────────────────────────────────────────
# EXPIRY TRACKER (partition-filtered)
# ─────────────────────────────────────────────────────────────
@app.route('/api/expiry', methods=['GET'])
@jwt_required()
def get_expiry():
    conn     = get_db()
    part     = _jwt_partition()
    identity = _get_identity()
    user_id  = identity['user_id']
    products = conn.execute(
        "SELECT * FROM products WHERE user_id=%s AND partition IN (%s, 'both')", (user_id, part)
    ).fetchall()
    conn.close()
    result   = {"expired": [], "within30": [], "within60": [], "within90": [], "safe": []}
    for p in products:
        days = expiry_days_left(p["expiry"])
        pd   = {**_product_out(p), "daysLeft": days}
        if days < 0:      result["expired"].append(pd)
        elif days <= 30:  result["within30"].append(pd)
        elif days <= 60:  result["within60"].append(pd)
        elif days <= 90:  result["within90"].append(pd)
        else:             result["safe"].append(pd)
    return jsonify({**result, "counts": {k: len(v) for k, v in result.items()}})


# ─────────────────────────────────────────────────────────────
# PARTITION INFO  — useful for debugging / frontend awareness
# ─────────────────────────────────────────────────────────────
@app.route('/api/partition-info', methods=['GET'])
@jwt_required()
def partition_info():
    """
    Returns info about which partition is active and record counts per partition.
    Useful for debugging and the settings UI.
    """
    conn = get_db()
    part = _jwt_partition()
    s    = _get_settings(conn)

    def count(table, p):
        return conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE partition=%s", (p,)
        ).fetchone()[0]
    def count_both(table):
        return conn.execute(
            f"SELECT COUNT(*) FROM {table} WHERE partition='both'"
        ).fetchone()[0]

    info = {
        "activePartition": part,
        "storeType":       s.get("store_type", "Retail Pharmacy"),
        "medicineCounts": {
            "shared":    count_both("products"),
            "wholesale": count("products", PARTITION_WS),
            "retail":    count("products", PARTITION_RT),
            "visible":   conn.execute(
                "SELECT COUNT(*) FROM products WHERE partition IN (%s, 'both')", (part,)
            ).fetchone()[0],
        },
        "stockInCounts": {
            "shared":    count_both("stock_ins"),
            "wholesale": count("stock_ins", PARTITION_WS),
            "retail":    count("stock_ins", PARTITION_RT),
        },
        "creditCounts": {
            "wholesale_shared":   count_both("credits"),
            "wholesale_specific": count("credits", PARTITION_WS),
            "retail_shared":      count_both("shop_credits"),
            "retail_specific":    count("shop_credits", PARTITION_RT),
        },
        "billCounts": {
            "wholesale": conn.execute(
                "SELECT COUNT(*) FROM bills WHERE bill_store_type='wholesale'"
            ).fetchone()[0],
            "retail": conn.execute(
                "SELECT COUNT(*) FROM bills WHERE bill_store_type!='wholesale'"
            ).fetchone()[0],
        },
    }
    conn.close()
    return jsonify(info)


# ─────────────────────────────────────────────────────────────
# PARALLEL QUERY HELPERS  (each grabs its own pool connection)
# All 8 helpers run simultaneously in get_state, cutting latency
# from  N × 200ms  →  ~200ms  (one Supabase round-trip total).
# ─────────────────────────────────────────────────────────────
def _pq_settings_user(user_id):
    c = get_db()
    try:
        s = c.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
        if not s:
            s = c.execute("SELECT * FROM settings ORDER BY id LIMIT 1").fetchone()
        u = c.execute("SELECT * FROM users WHERE id=%s", (user_id,)).fetchone()
        return (dict(s) if s else {}), (dict(u) if u else {})
    finally:
        c.close()

def _pq_categories(user_id):
    c = get_db()
    try:
        return [dict(r) for r in c.execute(
            "SELECT * FROM categories WHERE user_id=%s ORDER BY name", (user_id,)
        ).fetchall()]
    finally:
        c.close()

def _pq_products(user_id, part):
    c = get_db()
    try:
        return c.execute(
            "SELECT * FROM products WHERE user_id=%s AND partition IN (%s,'both') ORDER BY name",
            (user_id, part)
        ).fetchall()
    finally:
        c.close()

def _pq_bills(user_id, bill_w):
    c = get_db()
    try:
        bills_raw = c.execute(
            f"SELECT * FROM bills WHERE user_id=%s AND {bill_w} ORDER BY date DESC, created_at DESC LIMIT 500",
            (user_id,)
        ).fetchall()
        all_items = {}
        if bills_raw:
            ids = [b['id'] for b in bills_raw]
            ph  = ','.join(['%s'] * len(ids))
            for it in c.execute(
                f"SELECT * FROM bill_items WHERE bill_id IN ({ph})", ids
            ).fetchall():
                all_items.setdefault(it['bill_id'], []).append(_bill_item_out(it))
        return bills_raw, all_items
    finally:
        c.close()

def _pq_stock_ins(user_id, part):
    c = get_db()
    try:
        return c.execute(
            "SELECT * FROM stock_ins WHERE user_id=%s AND partition IN (%s,'both') ORDER BY date DESC LIMIT 300",
            (user_id, part)
        ).fetchall()
    finally:
        c.close()

def _pq_credits(user_id):
    c = get_db()
    try:
        return c.execute(
            "SELECT * FROM credits WHERE user_id=%s AND partition IN ('wholesale','both') ORDER BY date DESC",
            (user_id,)
        ).fetchall()
    finally:
        c.close()

def _pq_shop_credits(user_id):
    c = get_db()
    try:
        return c.execute(
            "SELECT * FROM shop_credits WHERE user_id=%s AND partition IN ('retail','both') ORDER BY bill_date DESC",
            (user_id,)
        ).fetchall()
    finally:
        c.close()

def _pq_resets(user_id):
    c = get_db()
    try:
        return c.execute(
            "SELECT store_type_key, reset_date FROM dashboard_resets WHERE user_id=%s", (user_id,)
        ).fetchall()
    finally:
        c.close()


def _pq_purchase_records(user_id):
    c = get_db()
    try:
        return c.execute(
            "SELECT * FROM purchase_records WHERE user_id=%s ORDER BY date DESC LIMIT 200",
            (user_id,)
        ).fetchall()
    finally:
        c.close()


def _purchase_record_out(r):
    return {
        "id":               r["id"],
        "date":             r["date"]             or "",
        "medicineName":     r["medicine_name"]    or "",
        "qty":              r["qty"]              or 0,
        "qtyUnit":          r["qty_unit"]         or "Box",
        "amountPaid":       r["amount_paid"]      or 0,
        "partyName":        r["party_name"]       or "",
        "partyType":        r["party_type"]       or "Supplier",
        "orderNo":          r["order_no"]         or "",
        "expectedDelivery": r["expected_delivery"] or "",
        "deliveryStatus":   r["delivery_status"]  or "Pending",
        "notes":            r["notes"]            or "",
    }


def _build_bill_fast(b, all_bill_items):
    return {
        "id":              b["id"],
        "billNo":          b["bill_no"]          or "",
        "date":            b["date"]             or "",
        "customer":        b["customer"]         or "",
        "phone":           b["phone"]            or "",
        "doctor":          b["doctor"]           or "",
        "rx":              b["rx"]               or "",
        "paymentMode":     b["payment_mode"]     or "Cash",
        "notes":           b["notes"]            or "",
        "subtotal":        b["subtotal"]         or 0,
        "totalDiscount":   b["total_discount"]   or 0,
        "totalGst":        b["total_gst"]        or 0,
        "roundOff":        b["round_off"]        or 0,
        "grandTotal":      b["grand_total"]      or 0,
        "billStoreType":   b["bill_store_type"]  or "retail",
        "wsSupplier":      b["ws_supplier"]      or "",
        "wsOwner":         b["ws_owner"]         or "",
        "wsGstin":         b["ws_gstin"]         or "",
        "shopName":        b["shop_name"]        or "",
        "shopkeeperGstin": b["shopkeeper_gstin"] or "",
        "rtShop":          b["rt_shop"]          or "",
        "rtOwner":         b["rt_owner"]         or "",
        "rtGstin":         b["rt_gstin"]         or "",
        "rtLicense":       b["rt_license"]       or "",
        "rtEmail":         b["rt_email"]         or "",
        "rtPhone":         b["rt_phone"]         or "",
        "items":           all_bill_items.get(b["id"], []),
    }

def _build_si(r):
    return {
        "id":          r["id"],
        "date":        r["date"]         or "",
        "productId":   r["product_id"]   or "",
        "productName": r["product_name"] or "",
        "qty":         r["qty"]          or 0,
        "price":       r["price"]        or 0,
        "batch":       r["batch"]        or "",
        "expiry":      r["expiry"]       or "",
        "supplier":    r["supplier"]     or "",
        "invoiceNo":   r["invoice_no"]   or "",
        "notes":       r["notes"]        or "",
        "partition":   r["partition"] if "partition" in r.keys() else PARTITION_BOTH,
    }


# ─────────────────────────────────────────────────────────────
# STATE — full STATE sync (partition-aware, parallel queries)
# ─────────────────────────────────────────────────────────────
@app.route('/api/state', methods=['GET'])
@jwt_required()
def get_state():
    """
    Return the entire app STATE filtered to the CURRENT partition.
    All 8 DB queries run in PARALLEL via ThreadPoolExecutor so the
    total latency is ~200ms (one Supabase RTT) instead of N×200ms.
    """
    identity = _get_identity()
    user_id  = identity['user_id']
    part     = _jwt_partition()
    bill_w   = _bill_type_filter(part)

    # ── Fire all 8 queries simultaneously ────────────────────────────────────
    with concurrent.futures.ThreadPoolExecutor(max_workers=9) as pool:
        f_su  = pool.submit(_pq_settings_user,    user_id)
        f_cat = pool.submit(_pq_categories,       user_id)
        f_pro = pool.submit(_pq_products,         user_id, part)
        f_bil = pool.submit(_pq_bills,            user_id, bill_w)
        f_si  = pool.submit(_pq_stock_ins,        user_id, part)
        f_cr  = pool.submit(_pq_credits,          user_id)
        f_sc  = pool.submit(_pq_shop_credits,     user_id)
        f_res = pool.submit(_pq_resets,           user_id)
        f_pr  = pool.submit(_pq_purchase_records, user_id)

    # ── Collect results (all already done by now) ─────────────────────────────
    s_raw, user        = f_su.result()
    categories         = f_cat.result()
    prods_raw          = f_pro.result()
    bills_raw, all_bi  = f_bil.result()
    si_raw             = f_si.result()
    credits_raw        = f_cr.result()
    shop_raw           = f_sc.result()
    resets_raw         = f_res.result()
    pr_raw             = f_pr.result()

    # ── Assemble response ─────────────────────────────────────────────────────
    s               = _settings_out_with_user(s_raw, user)
    products        = [_product_out(r)     for r in prods_raw]
    bills           = [_build_bill_fast(b, all_bi) for b in bills_raw]
    stock_ins       = [_build_si(r)        for r in si_raw]
    credits         = [_credit_out(r)      for r in credits_raw]
    shop_credits    = [_shop_credit_out(r) for r in shop_raw]
    dashboard_resets = {r["store_type_key"]: r["reset_date"] for r in resets_raw}
    purchase_records = [_purchase_record_out(r) for r in pr_raw]

    return jsonify({
        "settings":        s,
        "categories":      categories,
        "products":        products,
        "bills":           bills,
        "stockIns":        stock_ins,
        "credits":         credits,
        "shopCredits":     shop_credits,
        "nextBillNo":      s.get("nextBillNo", 1),
        "dashboardResets": dashboard_resets,
        "purchaseRecords": purchase_records,
    })


@app.route('/api/state', methods=['POST'])
@jwt_required()
@require_json
def save_state():
    """
    Partition-aware full-replace sync — called by app.js saveState().

    Strategy per data type:
    ─────────────────────────────────────────────────────────────
    categories   : Clear all + reinsert (globally shared)
    products     : Upsert payload items, preserving existing
                   'both' partition; delete current-partition items
                   that were removed from the frontend.
    stock_ins    : Same as products
    bills        : Replace only current partition's bill_store_type bills
    credits      : Clear all + reinsert (naturally wholesale-only table)
    shop_credits : Clear all + reinsert (naturally retail-only table)
    ─────────────────────────────────────────────────────────────
    """
    identity = _get_identity()
    user_id  = identity['user_id']
    data = request.get_json()
    conn = get_db()
    part = _jwt_partition()
    bill_type = PARTITION_WS if part == PARTITION_WS else PARTITION_RT

    # Settings are saved via PUT /api/settings — not here.
    # Updating settings inside save_state causes row-level lock contention
    # which produces the "QueryCanceled: statement timeout" 500 errors.
    # Only next_bill_no is synced here as it changes on every bill.
    s = data.get("settings", {})
    _nbn = int(data.get("nextBillNo", s.get("nextBillNo", 1)))
    try:
        conn.execute(
            "UPDATE settings SET next_bill_no=%s WHERE user_id=%s",
            (_nbn, user_id)
        )
        conn.commit()
    except Exception as _e:
        app.logger.warning(f"save_state: next_bill_no update skipped: {_e}")
        try: conn.rollback()
        except Exception: pass


    # ── SAFETY GUARD — never destructively wipe when frontend sends empty payload
    # Prevents accidental data loss if saveState fires before loadState completes.
    has_payload = bool(
        data.get("products") or data.get("categories") or data.get("bills") or
        data.get("stockIns") or data.get("credits") or data.get("shopCredits")
    )

    # ── Data sync — wrapped so connection always returns to pool ─────────────
    try:
        existing_prod_parts = {
            r["id"]: r["partition"]
            for r in conn.execute("SELECT id, partition FROM products WHERE user_id=%s", (user_id,)).fetchall()
        }
        existing_si_parts = {
            r["id"]: r["partition"]
            for r in conn.execute("SELECT id, partition FROM stock_ins WHERE user_id=%s", (user_id,)).fetchall()
        }
        payload_prod_ids = {p["id"]  for p in data.get("products", [])}
        payload_si_ids   = {si["id"] for si in data.get("stockIns", [])}

        # ── FK-SAFE DELETE ORDER: stock_ins → products → categories ──────────────
        # stock_ins.product_id references products(id)
        # products.category    references categories(id)
        # Deleting in reverse dependency order prevents ForeignKeyViolation.

        if has_payload:
            # 1. Delete removed stock_ins first (they reference products)
            if payload_si_ids:
                ph = ",".join(["%s"] * len(payload_si_ids))
                conn.execute(
                    f"DELETE FROM stock_ins WHERE user_id=%s AND partition=%s AND id NOT IN ({ph})",
                    [user_id, part] + list(payload_si_ids)
                )
            else:
                conn.execute("DELETE FROM stock_ins WHERE user_id=%s AND partition=%s", (user_id, part))

            # 2. Delete removed products next (they reference categories)
            if payload_prod_ids:
                ph = ",".join(["%s"] * len(payload_prod_ids))
                conn.execute(
                    f"DELETE FROM products WHERE user_id=%s AND partition=%s AND id NOT IN ({ph})",
                    [user_id, part] + list(payload_prod_ids)
                )
            else:
                conn.execute("DELETE FROM products WHERE user_id=%s AND partition=%s", (user_id, part))

            # 3. Categories: handled below with upsert + delete-removed-only.
            #    Do NOT do a full DELETE here — it triggers FK ON DELETE SET NULL
            #    across every products row, causing row-level lock contention and
            #    statement-timeout errors on Supabase free tier.
            pass

        # ── Categories — upsert payload items, delete only removed ones ──────────
        # This avoids the full-table DELETE that previously caused lock contention
        # timeouts on Supabase (FK cascade to products was too slow under 3s limit).
        payload_cat_ids = {c["id"] for c in data.get("categories", [])}
        for c in data.get("categories", []):
            conn.execute(
                'INSERT INTO categories(id,name,"desc",user_id) VALUES(%s,%s,%s,%s) '
                'ON CONFLICT (id) DO UPDATE SET name=EXCLUDED.name,"desc"=EXCLUDED."desc",user_id=EXCLUDED.user_id',
                (c["id"], c.get("name",""), c.get("desc",""), user_id))
        # Only delete categories the user explicitly removed (not in current payload)
        if has_payload:
            if payload_cat_ids:
                ph_cat = ",".join(["%s"] * len(payload_cat_ids))
                conn.execute(
                    f"DELETE FROM categories WHERE user_id=%s AND id NOT IN ({ph_cat})",
                    [user_id] + list(payload_cat_ids)
                )
            else:
                # Empty category list in payload = user cleared all categories
                conn.execute("DELETE FROM categories WHERE user_id=%s", (user_id,))

        # ── MEDICINE DB — Products upsert (categories rows exist again now) ───────
        for p in data.get("products", []):
            pid = p["id"]
            existing_p = existing_prod_parts.get(pid)
            final_part = PARTITION_BOTH if existing_p == PARTITION_BOTH else part
            conn.execute("""
                INSERT INTO products
                  (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,"desc",partition,user_id,selling_price,
                   pieces_per_strip,strips_per_box,purchase_unit)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                  purchase=EXCLUDED.purchase, sale=EXCLUDED.sale, gst=EXCLUDED.gst,
                  stock=EXCLUDED.stock, min_stock=EXCLUDED.min_stock, sku=EXCLUDED.sku,
                  expiry=EXCLUDED.expiry, brand=EXCLUDED.brand, hsn=EXCLUDED.hsn,
                  "desc"=EXCLUDED."desc", partition=EXCLUDED.partition, user_id=EXCLUDED.user_id,
                  selling_price=EXCLUDED.selling_price,
                  pieces_per_strip=EXCLUDED.pieces_per_strip,
                  strips_per_box=EXCLUDED.strips_per_box,
                  purchase_unit=EXCLUDED.purchase_unit
            """, (
                pid, p.get("name",""), p.get("category",""), p.get("unit","Tablet"),
                float(p.get("purchase",0)), float(p.get("sale",0)), float(p.get("gst",12)),
                int(p.get("stock",0)), int(p.get("minStock",10)),
                p.get("sku",""), p.get("expiry",""), p.get("brand",""),
                p.get("hsn",""), p.get("desc",""),
                final_part, user_id,
                float(p.get("sellingPrice", 0)),
                int(p.get("piecesPerStrip", 10) or 10),
                int(p.get("stripsPerBox",   10) or 10),
                (p.get("purchaseUnit", "box") or "box").lower(),
            ))

        # ── MEDICINE DB — Stock-ins upsert (products rows exist again now) ────────
        for si in data.get("stockIns", []):
            sid = si["id"]
            existing_sp = existing_si_parts.get(sid)
            final_part  = PARTITION_BOTH if existing_sp == PARTITION_BOTH else part
            conn.execute("""
                INSERT INTO stock_ins
                  (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition,user_id)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  date=EXCLUDED.date, product_id=EXCLUDED.product_id,
                  product_name=EXCLUDED.product_name, qty=EXCLUDED.qty,
                  price=EXCLUDED.price, batch=EXCLUDED.batch, expiry=EXCLUDED.expiry,
                  supplier=EXCLUDED.supplier, invoice_no=EXCLUDED.invoice_no,
                  notes=EXCLUDED.notes, partition=EXCLUDED.partition, user_id=EXCLUDED.user_id
            """, (
                sid, si.get("date",""), si.get("productId",""), si.get("productName",""),
                int(si.get("qty",0)), float(si.get("price",0)),
                si.get("batch",""), si.get("expiry",""), si.get("supplier",""),
                si.get("invoiceNo",""), si.get("notes",""),
                final_part, user_id,
            ))

        # ── SALES HISTORY DB — Bills (replace current partition's type) ─
        payload_bill_ids = {b["id"] for b in data.get("bills", [])}

        if bill_type == PARTITION_WS:
            btype_filter = "bill_store_type = 'wholesale'"
        else:
            btype_filter = "bill_store_type != 'wholesale'"

        if payload_bill_ids:
            ph = ",".join(["%s"] * len(payload_bill_ids))
            conn.execute(f"""
                DELETE FROM bill_items WHERE bill_id IN (
                    SELECT id FROM bills WHERE user_id=%s AND {btype_filter} AND id NOT IN ({ph})
                )""", [user_id] + list(payload_bill_ids))
            conn.execute(
                f"DELETE FROM bills WHERE user_id=%s AND {btype_filter} AND id NOT IN ({ph})",
                [user_id] + list(payload_bill_ids)
            )
        else:
            conn.execute(f"DELETE FROM bill_items WHERE bill_id IN (SELECT id FROM bills WHERE user_id=%s AND {btype_filter})", (user_id,))
            conn.execute(f"DELETE FROM bills WHERE user_id=%s AND {btype_filter}", (user_id,))

        for b in data.get("bills", []):
            # Only sync bills belonging to current partition
            b_type = b.get("billStoreType", "retail")
            if (part == PARTITION_WS and b_type != "wholesale") or \
               (part == PARTITION_RT and b_type == "wholesale"):
                continue

            conn.execute("""
                INSERT INTO bills
                  (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
                   subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
                   ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
                   rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone,user_id)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  bill_no=EXCLUDED.bill_no, date=EXCLUDED.date, customer=EXCLUDED.customer,
                  phone=EXCLUDED.phone, doctor=EXCLUDED.doctor, rx=EXCLUDED.rx,
                  payment_mode=EXCLUDED.payment_mode, notes=EXCLUDED.notes,
                  subtotal=EXCLUDED.subtotal, total_discount=EXCLUDED.total_discount,
                  total_gst=EXCLUDED.total_gst, round_off=EXCLUDED.round_off,
                  grand_total=EXCLUDED.grand_total, bill_store_type=EXCLUDED.bill_store_type,
                  ws_supplier=EXCLUDED.ws_supplier, ws_owner=EXCLUDED.ws_owner,
                  ws_gstin=EXCLUDED.ws_gstin, shop_name=EXCLUDED.shop_name,
                  shopkeeper_gstin=EXCLUDED.shopkeeper_gstin, rt_shop=EXCLUDED.rt_shop,
                  rt_owner=EXCLUDED.rt_owner, rt_gstin=EXCLUDED.rt_gstin,
                  rt_license=EXCLUDED.rt_license, rt_email=EXCLUDED.rt_email,
                  rt_phone=EXCLUDED.rt_phone, user_id=EXCLUDED.user_id
            """, (
                b["id"], b.get("billNo",""), b.get("date",""),
                b.get("customer",""), b.get("phone",""), b.get("doctor",""), b.get("rx",""),
                b.get("paymentMode","Cash"), b.get("notes",""),
                float(b.get("subtotal",0)), float(b.get("totalDiscount",0)),
                float(b.get("totalGst",0)), float(b.get("roundOff",0)), float(b.get("grandTotal",0)),
                b.get("billStoreType","retail"),
                b.get("wsSupplier",""), b.get("wsOwner",""), b.get("wsGstin",""),
                b.get("shopName",""), b.get("shopkeeperGstin",""),
                b.get("rtShop",""), b.get("rtOwner",""), b.get("rtGstin",""),
                b.get("rtLicense",""), b.get("rtEmail",""), b.get("rtPhone",""),
                user_id,
            ))
            for it in b.get("items", []):
                ga, lt = _calc_item(
                    float(it.get("unitPrice",0)), float(it.get("qty",0)),
                    float(it.get("discount",0)), float(it.get("gstRate",0)))
                conn.execute("""
                    INSERT INTO bill_items
                      (id,bill_id,product_id,name,category,unit,qty,unit_price,
                       discount,gst_rate,gst_amt,line_total,
                       unit_type,display_qty,qty_in_pieces,amount_before_tax,
                       mrp_per_box,selling_price_per_box,strips_per_box,pieces_per_strip)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (id) DO UPDATE SET
                      bill_id=EXCLUDED.bill_id, product_id=EXCLUDED.product_id,
                      name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                      qty=EXCLUDED.qty, unit_price=EXCLUDED.unit_price,
                      discount=EXCLUDED.discount, gst_rate=EXCLUDED.gst_rate,
                      gst_amt=EXCLUDED.gst_amt, line_total=EXCLUDED.line_total,
                      unit_type=EXCLUDED.unit_type, display_qty=EXCLUDED.display_qty,
                      qty_in_pieces=EXCLUDED.qty_in_pieces, amount_before_tax=EXCLUDED.amount_before_tax,
                      mrp_per_box=EXCLUDED.mrp_per_box, selling_price_per_box=EXCLUDED.selling_price_per_box,
                      strips_per_box=EXCLUDED.strips_per_box, pieces_per_strip=EXCLUDED.pieces_per_strip
                """, (
                    it.get("id") or uid(), b["id"],
                    it.get("productId",""), it.get("name",""),
                    it.get("category",""), it.get("unit",""),
                    float(it.get("qty",0)), float(it.get("unitPrice",0)),
                    float(it.get("discount",0)), float(it.get("gstRate",0)),
                    float(it.get("gstAmt", ga)), float(it.get("lineTotal", lt)),
                    (it.get("unitType") or "strip").lower(),
                    float(it.get("displayQty") or it.get("qty",0)),
                    float(it.get("qtyInPieces") or it.get("qty",0)),
                    float(it.get("amountBeforeTax",0)),
                    float(it.get("mrpPerBox",0)),
                    float(it.get("sellingPricePerBox",0)),
                    int(it.get("stripsPerBox",10) or 10),
                    int(it.get("piecesPerStrip",10) or 10),
                ))

        # ── CREDIT DB — Wholesale credits (user-isolated full replace) ─
        conn.execute("DELETE FROM credits WHERE user_id=%s", (user_id,))
        for c in data.get("credits", []):
            conn.execute("""
                INSERT INTO credits
                  (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition,user_id)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  date=EXCLUDED.date, shop_name=EXCLUDED.shop_name,
                  shopkeeper_name=EXCLUDED.shopkeeper_name, phone=EXCLUDED.phone,
                  for_item=EXCLUDED.for_item, amount=EXCLUDED.amount,
                  method=EXCLUDED.method, status=EXCLUDED.status,
                  partition=EXCLUDED.partition, user_id=EXCLUDED.user_id
            """, (
                c["id"], c.get("date",""), c.get("shopName",""), c.get("shopkeeperName",""),
                c.get("phone",""), c.get("forItem",""),
                float(c.get("amount",0)),
                c.get("method","Cash"), c.get("status","Pending"),
                c.get("partition", PARTITION_WS), user_id,
            ))

        # ── CREDIT DB — Retail shop_credits (user-isolated full replace) ─
        conn.execute("DELETE FROM shop_credits WHERE user_id=%s", (user_id,))
        for sc in data.get("shopCredits", []):
            pending = sc.get("pending", max(0, sc.get("totalPurchase",0) - sc.get("paid",0)))
            conn.execute("""
                INSERT INTO shop_credits
                  (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
                   payment_mode,pending,last_purchase_date,bill_date,status,partition,user_id)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  supplier_id=EXCLUDED.supplier_id, supplier_name=EXCLUDED.supplier_name,
                  owner_name=EXCLUDED.owner_name, total_purchase=EXCLUDED.total_purchase,
                  paid=EXCLUDED.paid, payment_mode=EXCLUDED.payment_mode,
                  pending=EXCLUDED.pending, last_purchase_date=EXCLUDED.last_purchase_date,
                  bill_date=EXCLUDED.bill_date, status=EXCLUDED.status,
                  partition=EXCLUDED.partition, user_id=EXCLUDED.user_id
            """, (
                sc["id"], sc.get("supplierId",""), sc.get("supplierName",""), sc.get("ownerName",""),
                float(sc.get("totalPurchase",0)), float(sc.get("paid",0)),
                sc.get("paymentMode","Cash"), float(pending),
                sc.get("lastPurchaseDate",""), sc.get("billDate",""),
                sc.get("status","Pending"),
                sc.get("partition", PARTITION_RT), user_id,
            ))

        # ── PURCHASE RECORDS — full replace per user (wholesale personal ledger) ──
        conn.execute("DELETE FROM purchase_records WHERE user_id=%s", (user_id,))
        for pr in data.get("purchaseRecords", []):
            conn.execute("""
                INSERT INTO purchase_records
                  (id,date,medicine_name,qty,qty_unit,amount_paid,party_name,
                   party_type,order_no,expected_delivery,delivery_status,notes,user_id)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  date=EXCLUDED.date, medicine_name=EXCLUDED.medicine_name,
                  qty=EXCLUDED.qty, qty_unit=EXCLUDED.qty_unit,
                  amount_paid=EXCLUDED.amount_paid, party_name=EXCLUDED.party_name,
                  party_type=EXCLUDED.party_type, order_no=EXCLUDED.order_no,
                  expected_delivery=EXCLUDED.expected_delivery,
                  delivery_status=EXCLUDED.delivery_status, notes=EXCLUDED.notes,
                  user_id=EXCLUDED.user_id
            """, (
                pr["id"], pr.get("date",""), pr.get("medicineName",""),
                float(pr.get("qty",0)), pr.get("qtyUnit","Box"),
                float(pr.get("amountPaid",0)), pr.get("partyName",""),
                pr.get("partyType","Supplier"), pr.get("orderNo",""),
                pr.get("expectedDelivery",""), pr.get("deliveryStatus","Pending"),
                pr.get("notes",""), user_id,
            ))

        conn.commit()
    except Exception as _sync_err:
        app.logger.error(f"save_state error: {_sync_err}")
        try: conn.rollback()
        except Exception: pass
        return jsonify({"error": "Database sync error", "detail": str(_sync_err)}), 500
    finally:
        conn.close()

    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# IMPORT  (migrate from localStorage STATE JSON)
# All imported records are tagged partition='both' so they are
# visible regardless of which pharmacy mode is active.
# ─────────────────────────────────────────────────────────────
@app.route('/api/import', methods=['POST'])
@jwt_required()
@require_json
def import_data():
    """
    One-shot migration from the original localStorage STATE object.
    POST the JSON from localStorage key 'pharmacare_v2'.
    Safe to call multiple times — uses INSERT OR IGNORE.
    All imported records are tagged partition='both' (visible in all modes).
    """
    data = request.get_json()
    conn = get_db()

    identity = _get_identity()
    user_id  = identity['user_id']
    # Settings (per-user)
    s = data.get("settings", {})
    existing_s = conn.execute("SELECT id FROM settings WHERE user_id=%s", (user_id,)).fetchone()
    if not existing_s:
        conn.execute("INSERT INTO settings (user_id, store_type) VALUES (%s,%s)",
                     (user_id, identity.get('pharmacy_type','Retail Pharmacy')))
    conn.execute("""
        UPDATE settings SET
          store_name=%s, store_type=%s, address=%s, phone=%s, email=%s,
          license_no=%s, gstin=%s, default_gst=%s, currency=%s,
          low_stock_threshold=%s, expiry_alert_days=%s,
          wholesaler=%s, owner_name=%s, wholesaler_id=%s,
          shop_name=%s, retailer_owner=%s,
          wholesale_upi_qr=%s, retail_upi_qr=%s, next_bill_no=%s
        WHERE user_id=%s
    """, (
        s.get("storeName",         "My Pharmacy"),
        identity.get("pharmacy_type","Retail Pharmacy"),
        s.get("address",           ""),
        s.get("phone",             ""),
        s.get("email",             ""),
        identity.get("drug_license",""),
        identity.get("gstin",      ""),
        float(s.get("defaultGst",  12)),
        s.get("currency",          "₹"),
        int(s.get("lowStockThreshold", 10)),
        int(s.get("expiryAlertDays",   90)),
        s.get("supplierName",      ""),
        s.get("wholesaler",        ""),
        s.get("ownerName",         ""),
        s.get("wholesalerId",      ""),
        s.get("shopName",          ""),
        s.get("retailerOwner",     ""),
        s.get("wholesaleUpiQr",    ""),
        s.get("retailUpiQr",       ""),
        int(data.get("nextBillNo", s.get("nextBillNo", 1))),
        user_id,
    ))

    for c in data.get("categories", []):
        conn.execute(
            'INSERT INTO categories(id,name,"desc") VALUES(%s,%s,%s) ON CONFLICT (id) DO NOTHING',
            (c["id"], c["name"], c.get("desc", ""))
        )

    # All migrated products → partition='both' (visible in all modes)
    for p in data.get("products", []):
        conn.execute("""
            INSERT INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,hsn,"desc",partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO UPDATE SET
              name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
              purchase=EXCLUDED.purchase, sale=EXCLUDED.sale, gst=EXCLUDED.gst,
              stock=EXCLUDED.stock, min_stock=EXCLUDED.min_stock, sku=EXCLUDED.sku,
              expiry=EXCLUDED.expiry, brand=EXCLUDED.brand, hsn=EXCLUDED.hsn,
              "desc"=EXCLUDED."desc", partition=EXCLUDED.partition
        """, (
            p["id"], p["name"],
            p.get("category", ""),
            p.get("unit",     "Tablet"),
            float(p.get("purchase",  0)),
            float(p.get("sale",      0)),
            float(p.get("gst",      12)),
            int(p.get("stock",       0)),
            int(p.get("minStock",   10)),
            p.get("sku",    ""),
            p.get("expiry", ""),
            p.get("brand",  ""),
            p.get("hsn",    ""),
            p.get("desc",   ""),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    # All migrated stock-ins → partition='both'
    for si in data.get("stockIns", []):
        conn.execute("""
            INSERT INTO stock_ins
              (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,notes,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO UPDATE SET
              date=EXCLUDED.date, product_id=EXCLUDED.product_id,
              product_name=EXCLUDED.product_name, qty=EXCLUDED.qty,
              price=EXCLUDED.price, batch=EXCLUDED.batch, expiry=EXCLUDED.expiry,
              supplier=EXCLUDED.supplier, invoice_no=EXCLUDED.invoice_no,
              notes=EXCLUDED.notes, partition=EXCLUDED.partition
        """, (
            si["id"], si.get("date", ""),
            si.get("productId", ""), si.get("productName", ""),
            int(si.get("qty", 0)), float(si.get("price", 0)),
            si.get("batch", ""), si.get("expiry", ""),
            si.get("supplier", ""), si.get("invoiceNo", ""), si.get("notes", ""),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    for b in data.get("bills", []):
        conn.execute("""
            INSERT INTO bills
              (id,bill_no,date,customer,phone,doctor,rx,payment_mode,notes,
               subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type,
               ws_supplier,ws_owner,ws_gstin,shop_name,shopkeeper_gstin,
               rt_shop,rt_owner,rt_gstin,rt_license,rt_email,rt_phone)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO UPDATE SET
              bill_no=EXCLUDED.bill_no, date=EXCLUDED.date, customer=EXCLUDED.customer,
              phone=EXCLUDED.phone, doctor=EXCLUDED.doctor, rx=EXCLUDED.rx,
              payment_mode=EXCLUDED.payment_mode, notes=EXCLUDED.notes,
              subtotal=EXCLUDED.subtotal, total_discount=EXCLUDED.total_discount,
              total_gst=EXCLUDED.total_gst, round_off=EXCLUDED.round_off,
              grand_total=EXCLUDED.grand_total, bill_store_type=EXCLUDED.bill_store_type,
              ws_supplier=EXCLUDED.ws_supplier, ws_owner=EXCLUDED.ws_owner,
              ws_gstin=EXCLUDED.ws_gstin, shop_name=EXCLUDED.shop_name,
              shopkeeper_gstin=EXCLUDED.shopkeeper_gstin, rt_shop=EXCLUDED.rt_shop,
              rt_owner=EXCLUDED.rt_owner, rt_gstin=EXCLUDED.rt_gstin,
              rt_license=EXCLUDED.rt_license, rt_email=EXCLUDED.rt_email,
              rt_phone=EXCLUDED.rt_phone
        """, (
            b["id"], b.get("billNo", ""), b.get("date", ""),
            b.get("customer", ""), b.get("phone", ""),
            b.get("doctor", ""), b.get("rx", ""),
            b.get("paymentMode", "Cash"), b.get("notes", ""),
            float(b.get("subtotal",      0)),
            float(b.get("totalDiscount", 0)),
            float(b.get("totalGst",      0)),
            float(b.get("roundOff",      0)),
            float(b.get("grandTotal",    0)),
            b.get("billStoreType", "retail"),
            b.get("wsSupplier", ""), b.get("wsOwner", ""),
            b.get("wsGstin", ""),   b.get("shopName", ""),
            b.get("shopkeeperGstin", ""),
            b.get("rtShop", ""),    b.get("rtOwner", ""),
            b.get("rtGstin", ""),   b.get("rtLicense", ""),
            b.get("rtEmail", ""),   b.get("rtPhone", ""),
        ))
        for it in b.get("items", []):
            gst_amt, line_total = _calc_item(
                float(it.get("unitPrice", 0)), float(it.get("qty", 0)),
                float(it.get("discount",  0)), float(it.get("gstRate", 0)),
            )
            conn.execute("""
                INSERT INTO bill_items
                  (id,bill_id,product_id,name,category,unit,qty,unit_price,
                   discount,gst_rate,gst_amt,line_total)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  bill_id=EXCLUDED.bill_id, product_id=EXCLUDED.product_id,
                  name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                  qty=EXCLUDED.qty, unit_price=EXCLUDED.unit_price,
                  discount=EXCLUDED.discount, gst_rate=EXCLUDED.gst_rate,
                  gst_amt=EXCLUDED.gst_amt, line_total=EXCLUDED.line_total
            """, (
                it.get("id") or uid(), b["id"],
                it.get("productId", ""), it.get("name", ""),
                it.get("category",  ""), it.get("unit",  ""),
                float(it.get("qty",       0)),
                float(it.get("unitPrice", 0)),
                float(it.get("discount",  0)),
                float(it.get("gstRate",   0)),
                float(it.get("gstAmt",    gst_amt)),
                float(it.get("lineTotal", line_total)),
            ))

    # Migrated wholesale credits → partition='both'
    for c in data.get("credits", []):
        conn.execute("""
            INSERT INTO credits
              (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition,user_id)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (
            c["id"], c.get("date", ""),
            c.get("shopName", ""), c.get("shopkeeperName", ""),
            c.get("phone", ""), c.get("forItem", ""),
            float(c.get("amount", 0)),
            c.get("method", "Cash"), c.get("status", "Pending"),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    # Migrated retail shop_credits → partition='both'
    for sc in data.get("shopCredits", []):
        pending = sc.get("pending", max(0, sc.get("totalPurchase", 0) - sc.get("paid", 0)))
        conn.execute("""
            INSERT INTO shop_credits
              (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
               payment_mode,pending,last_purchase_date,bill_date,status,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (
            sc["id"],
            sc.get("supplierId", ""), sc.get("supplierName", ""),
            sc.get("ownerName", ""),
            float(sc.get("totalPurchase", 0)),
            float(sc.get("paid",          0)),
            sc.get("paymentMode", "Cash"),
            float(pending),
            sc.get("lastPurchaseDate", ""),
            sc.get("billDate", ""),
            sc.get("status", "Pending"),
            PARTITION_BOTH,     # ← visible to all modes
        ))

    conn.commit()
    counts = {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
              for t in ["categories","products","bills","stock_ins","credits","shop_credits"]}
    conn.close()
    return jsonify({"ok": True, "message": "Data imported successfully (partition=both)", "counts": counts})


# ─────────────────────────────────────────────────────────────
# FULL BACKUP EXPORT  (all partitions — for complete backup)
# ─────────────────────────────────────────────────────────────
@app.route('/api/export/backup', methods=['GET'])
@jwt_required()
def export_full_backup():
    """Return ALL data from every partition as a single JSON backup."""
    conn = get_db()
    s    = _get_settings(conn)

    categories = rows(conn.execute("SELECT * FROM categories ORDER BY name").fetchall())
    products   = [_product_out(r) for r in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]

    # All bills (both partitions)
    all_bills = []
    for b in conn.execute("SELECT * FROM bills ORDER BY date DESC").fetchall():
        items = rows(conn.execute("SELECT * FROM bill_items WHERE bill_id=%s", (b["id"],)).fetchall())
        bill_dict = dict(b)
        bill_dict["items"] = [_bill_item_out(i) for i in conn.execute(
            "SELECT * FROM bill_items WHERE bill_id=%s", (b["id"],)).fetchall()]
        all_bills.append({
            "id": b["id"], "billNo": b["bill_no"] or "", "date": b["date"] or "",
            "customer": b["customer"] or "", "phone": b["phone"] or "",
            "doctor": b["doctor"] or "", "rx": b["rx"] or "",
            "paymentMode": b["payment_mode"] or "Cash", "notes": b["notes"] or "",
            "subtotal": b["subtotal"] or 0, "totalDiscount": b["total_discount"] or 0,
            "totalGst": b["total_gst"] or 0, "roundOff": b["round_off"] or 0,
            "grandTotal": b["grand_total"] or 0,
            "billStoreType": b["bill_store_type"] or "retail",
            "wsSupplier": b["ws_supplier"] or "", "wsOwner": b["ws_owner"] or "",
            "wsGstin": b["ws_gstin"] or "", "shopName": b["shop_name"] or "",
            "shopkeeperGstin": b["shopkeeper_gstin"] or "",
            "rtShop": b["rt_shop"] or "", "rtOwner": b["rt_owner"] or "",
            "rtGstin": b["rt_gstin"] or "", "rtLicense": b["rt_license"] or "",
            "rtEmail": b["rt_email"] or "", "rtPhone": b["rt_phone"] or "",
            "items": [_bill_item_out(i) for i in conn.execute(
                "SELECT * FROM bill_items WHERE bill_id=%s", (b["id"],)).fetchall()],
        })

    stock_ins    = [{"id": r["id"], "date": r["date"] or "", "productId": r["product_id"] or "",
                     "productName": r["product_name"] or "", "qty": r["qty"] or 0,
                     "price": r["price"] or 0, "batch": r["batch"] or "",
                     "expiry": r["expiry"] or "", "supplier": r["supplier"] or "",
                     "invoiceNo": r["invoice_no"] or "", "notes": r["notes"] or "",
                     "partition": r["partition"] if "partition" in r.keys() else PARTITION_BOTH}
                    for r in conn.execute("SELECT * FROM stock_ins ORDER BY date DESC").fetchall()]
    credits      = [_credit_out(r) for r in conn.execute("SELECT * FROM credits ORDER BY date DESC").fetchall()]
    shop_credits = [_shop_credit_out(r) for r in conn.execute("SELECT * FROM shop_credits ORDER BY bill_date DESC").fetchall()]

    reset_rows = conn.execute("SELECT store_type_key, reset_date FROM dashboard_resets").fetchall()
    dashboard_resets = {r["store_type_key"]: r["reset_date"] for r in reset_rows}
    conn.close()

    return jsonify({
        "settings":        _settings_out(s),
        "categories":      categories,
        "products":        products,
        "bills":           all_bills,
        "stockIns":        stock_ins,
        "credits":         credits,
        "shopCredits":     shop_credits,
        "nextBillNo":      s.get("next_bill_no", 1),
        "dashboardResets": dashboard_resets,
        "_exportMeta": {
            "exportedAt": date.today().isoformat(),
            "source":     "PharmaCare Pro Backup",
            "version":    "2.0",
        }
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Medicines only  (CSV / XLSX pre-parsed to JSON)
# Partition assigned based on current pharmacy type in settings.
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/medicines', methods=['POST'])
@jwt_required()
@require_json
def import_medicines():
    """
    Import medicine records from a JSON array.
    Frontend parses CSV/XLSX to JSON, sends here.
    Expected body: { "medicines": [ { name, category, unit, purchase, sale, gst,
                                       stock, minStock, sku, expiry, brand, hsn, "desc" }, ... ] }
    Partition is set from the current pharmacy type in settings.
    """
    data     = request.get_json()
    medicines = data.get("medicines", [])
    if not medicines:
        return jsonify({"error": "No medicine data found in the uploaded file"}), 400

    conn = get_db()
    part = _jwt_partition()

    inserted = 0
    updated  = 0
    skipped  = 0
    errors   = []

    for m in medicines:
        name = (m.get("name") or "").strip()
        if not name:
            skipped += 1
            continue

        # Resolve category — accept name or id
        cat_raw = (m.get("category") or "").strip()
        cat_id  = None
        if cat_raw:
            # Try as ID first
            row_c = conn.execute("SELECT id FROM categories WHERE id=%s", (cat_raw,)).fetchone()
            if row_c:
                cat_id = row_c["id"]
            else:
                # Try as name (case-insensitive)
                row_c = conn.execute(
                    "SELECT id FROM categories WHERE LOWER(name)=%s", (cat_raw.lower(),)
                ).fetchone()
                if row_c:
                    cat_id = row_c["id"]
                else:
                    # Create the category
                    cat_id = uid()
                    conn.execute(
                        'INSERT INTO categories(id,name,"desc") VALUES(%s,%s,%s) ON CONFLICT (id) DO NOTHING',
                        (cat_id, cat_raw, "")
                    )

        pid = m.get("id") or uid()

        # Check if already exists
        existing = conn.execute("SELECT id FROM products WHERE id=%s", (pid,)).fetchone()
        if not existing:
            # Also check by name to avoid exact duplicates
            existing_by_name = conn.execute(
                "SELECT id FROM products WHERE LOWER(name)=%s AND partition=%s",
                (name.lower(), part)
            ).fetchone()
            if existing_by_name:
                pid = existing_by_name["id"]

        try:
            conn.execute("""
                INSERT INTO products
                  (id,name,category,unit,purchase,sale,gst,stock,min_stock,
                   sku,expiry,brand,hsn,"desc",partition)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                  purchase=EXCLUDED.purchase, sale=EXCLUDED.sale, gst=EXCLUDED.gst,
                  stock=EXCLUDED.stock, min_stock=EXCLUDED.min_stock, sku=EXCLUDED.sku,
                  expiry=EXCLUDED.expiry, brand=EXCLUDED.brand, hsn=EXCLUDED.hsn,
                  "desc"=EXCLUDED."desc", partition=EXCLUDED.partition
            """, (
                pid, name, cat_id,
                (m.get("unit") or m.get("form") or "Tablet"),
                float(m.get("purchase") or m.get("pur") or 0),
                float(m.get("sale") or m.get("mrp") or 0),
                float(m.get("gst") or 12),
                int(m.get("stock") or m.get("stk") or 0),
                int(m.get("minStock") or m.get("min_stock") or m.get("ms") or 10),
                (m.get("sku") or m.get("batch") or m.get("bat") or ""),
                (m.get("expiry") or m.get("exp") or ""),
                (m.get("brand") or m.get("manufacturer") or m.get("mfr") or ""),
                (m.get("hsn") or ""),
                (m.get("desc") or m.get("composition") or m.get("cmp") or ""),
                part,
            ))
            if existing:
                updated += 1
            else:
                inserted += 1
        except Exception as ex:
            errors.append({"name": name, "error": str(ex)})
            skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Medicines imported into '{part}' partition",
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "errors":  errors[:10],  # return first 10 errors only
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Sales History  (CSV / XLSX pre-parsed to JSON)
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/sales-history', methods=['POST'])
@jwt_required()
@require_json
def import_sales_history():
    """
    Import sales/bill records.
    Expected body: { "bills": [ { billNo, date, customer, phone, doctor,
                                   paymentMode, grandTotal, subtotal, ... }, ... ] }
    bill_store_type is set from current partition.
    """
    data  = request.get_json()
    bills = data.get("bills", [])
    if not bills:
        return jsonify({"error": "No sales data found in the uploaded file"}), 400

    conn      = get_db()
    part      = _jwt_partition()
    bill_type = PARTITION_WS if part == PARTITION_WS else PARTITION_RT

    inserted = 0
    skipped  = 0
    errors   = []

    for b in bills:
        bill_date = (b.get("date") or b.get("Date") or today_str())
        customer  = (b.get("customer") or b.get("Customer") or b.get("customerName") or "").strip()
        grand     = float(b.get("grandTotal") or b.get("grand_total") or b.get("GrandTotal") or 0)
        subtotal  = float(b.get("subtotal") or b.get("Subtotal") or grand)
        bid       = b.get("id") or uid()

        existing = conn.execute("SELECT id FROM bills WHERE id=%s", (bid,)).fetchone()
        if existing:
            skipped += 1
            continue

        # Auto-increment bill number
        s_row = conn.execute("SELECT * FROM settings WHERE user_id=%s", (user_id,)).fetchone()
        s = dict(s_row) if s_row else _get_settings(conn)
        n   = s.get("next_bill_no", 1)
        bn  = str(n).zfill(4)
        conn.execute("UPDATE settings SET next_bill_no=%s WHERE user_id=%s", (n + 1, user_id))

        try:
            conn.execute("""
                INSERT INTO bills
                  (id,bill_no,date,customer,phone,doctor,payment_mode,
                   subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO NOTHING
            """, (
                bid,
                b.get("billNo") or b.get("bill_no") or bn,
                bill_date, customer,
                (b.get("phone") or ""),
                (b.get("doctor") or ""),
                (b.get("paymentMode") or b.get("payment_mode") or b.get("PaymentMode") or "Cash"),
                subtotal,
                float(b.get("totalDiscount") or b.get("total_discount") or 0),
                float(b.get("totalGst") or b.get("total_gst") or 0),
                float(b.get("roundOff") or b.get("round_off") or 0),
                grand,
                bill_type,
            ))
            inserted += 1
        except Exception as ex:
            errors.append({"customer": customer, "error": str(ex)})
            skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Sales history imported into '{bill_type}' store type",
        "inserted": inserted, "skipped": skipped,
        "errors": errors[:10],
    })


# ─────────────────────────────────────────────────────────────
# IMPORT — Credits  (CSV / XLSX pre-parsed to JSON)
# ─────────────────────────────────────────────────────────────
@app.route('/api/import/credits', methods=['POST'])
@jwt_required()
@require_json
def import_credits():
    """
    Import credit records.
    Wholesale: credits table. Retail/Hospital/Medical/Ayurvedic: shop_credits table.
    """
    data    = request.get_json()
    credits_data = data.get("credits", [])
    if not credits_data:
        return jsonify({"error": "No credit data found in the uploaded file"}), 400

    conn = get_db()
    part = _jwt_partition()

    inserted = 0
    skipped  = 0
    errors   = []

    if part == PARTITION_WS:
        # Insert into wholesale credits table
        for c in credits_data:
            shop = (c.get("shopName") or c.get("shop_name") or c.get("ShopName") or "").strip()
            if not shop:
                skipped += 1
                continue
            cid = c.get("id") or uid()
            existing = conn.execute("SELECT id FROM credits WHERE id=%s", (cid,)).fetchone()
            if existing:
                skipped += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO credits
                      (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (id) DO NOTHING
                """, (
                    cid,
                    (c.get("date") or c.get("Date") or today_str()),
                    shop,
                    (c.get("shopkeeperName") or c.get("shopkeeper_name") or c.get("OwnerName") or ""),
                    (c.get("phone") or ""),
                    (c.get("forItem") or c.get("for_item") or c.get("Item") or ""),
                    float(c.get("amount") or c.get("Amount") or 0),
                    (c.get("method") or c.get("Method") or "Cash"),
                    (c.get("status") or "Pending"),
                    PARTITION_WS,
                ))
                inserted += 1
            except Exception as ex:
                errors.append({"shop": shop, "error": str(ex)})
                skipped += 1
    else:
        # Insert into retail shop_credits table
        for sc in credits_data:
            supplier = (sc.get("supplierName") or sc.get("supplier_name") or sc.get("SupplierName") or "").strip()
            if not supplier:
                skipped += 1
                continue
            scid = sc.get("id") or uid()
            existing = conn.execute("SELECT id FROM shop_credits WHERE id=%s", (scid,)).fetchone()
            if existing:
                skipped += 1
                continue
            try:
                total  = float(sc.get("totalPurchase") or sc.get("total_purchase") or 0)
                paid   = float(sc.get("paid") or 0)
                pending = float(sc.get("pending") or max(0, total - paid))
                conn.execute("""
                    INSERT INTO shop_credits
                      (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
                       payment_mode,pending,last_purchase_date,bill_date,status,partition)
                    VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (id) DO NOTHING
                """, (
                    scid,
                    (sc.get("supplierId") or sc.get("supplier_id") or ""),
                    supplier,
                    (sc.get("ownerName") or sc.get("owner_name") or ""),
                    total, paid,
                    (sc.get("paymentMode") or sc.get("payment_mode") or "Cash"),
                    pending,
                    (sc.get("lastPurchaseDate") or sc.get("last_purchase_date") or today_str()),
                    (sc.get("billDate") or sc.get("bill_date") or today_str()),
                    (sc.get("status") or ("Cleared" if pending <= 0 else "Pending")),
                    PARTITION_RT,
                ))
                inserted += 1
            except Exception as ex:
                errors.append({"supplier": supplier, "error": str(ex)})
                skipped += 1

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "partition": part,
        "message": f"Credits imported into '{part}' partition",
        "inserted": inserted, "skipped": skipped,
        "errors": errors[:10],
    })


# ─────────────────────────────────────────────────────────────
# SEED EXCEL MEDICINES  — loads med_query__1_.xlsx on first run
# All records tagged partition='both' (visible in ALL modes).
# Column mapping:
#   Column1.id   → id          Column1.name → name
#   Column1.cid  → category    Column1.form → unit
#   Column1.pur  → purchase    Column1.mrp  → sale
#   Column1.gst  → gst         Column1.stk  → stock
#   Column1.ms   → min_stock   Column1.bat  → sku (batch)
#   Column1.exp  → expiry      Column1.mfr  → brand
#   Column1.hsn  → hsn         Column1.cmp  → "desc"
# ─────────────────────────────────────────────────────────────
def seed_excel_medicines():
    """
    Load medicines from med_query__1_.xlsx into products table.
    Safe to call multiple times — skips entirely if already seeded.
    All records tagged partition='both' → visible in ALL pharmacy modes.
    """
    xlsx_path = os.path.join(BASE_DIR, 'med_query__1_.xlsx')
    if not os.path.exists(xlsx_path):
        print("  ℹ  med_query__1_.xlsx not found — skipping Excel medicine seed")
        return

    # ── FAST PATH: skip entirely if seed data already exists ─────────────────
    conn = get_db()
    existing = conn.execute("SELECT COUNT(*) FROM products WHERE user_id IS NULL").fetchone()[0]
    conn.close()
    if existing > 0:
        print(f"  ✓ Excel medicines already seeded ({existing} seed records) — skipping Excel parse")
        return
    # ─────────────────────────────────────────────────────────────────────────

    try:
        import csv, io
        # Try pandas first, fall back to openpyxl
        try:
            import pandas as pd
            df = pd.read_excel(xlsx_path)
            records = df.to_dict('records')
        except ImportError:
            from openpyxl import load_workbook
            wb  = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws  = wb.active
            hdr = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(hdr, row)))
    except Exception as e:
        print(f"  ✗ Could not read med_query__1_.xlsx: {e}")
        return

    if not records:
        print("  ℹ  med_query__1_.xlsx is empty — skipping")
        return

    conn = get_db()

    # ── Build a map of category_id → category_id (they're already IDs in the sheet)
    # We'll insert categories on-the-fly if they don't exist.
    # The sheet uses opaque hex IDs for categories (Column1.cid).
    # Map known cid values to human-readable names from the dataset.
    cid_name_map = {
        '5f122f32': 'Analgesics',
        '198ee715': 'Antibiotics',
        '70dfa1e7': 'Antacids & GI',
        '892b6604': 'Antihistamines',
        'e995823e': 'Vitamins & Supplements',
        'fa48a4f5': 'Antidiabetics',
        '15c1fd0a': 'Cardiovascular',
        'b4681e22': 'Syrups & Liquids',
        '9048aed2': 'Topical',
        '30473ab8': 'Respiratory',
        '6c8109f5': 'Neurological',
        '79c2f85f': 'Hormonal',
        'b299e67d': 'Eye Care',
        'd8c361a7': 'Dental & Oral',
        '41b2800f': 'Surgical & Consumables',
    }

    inserted = 0
    skipped  = 0

    for rec in records:
        # Column names as they appear in the xlsx (with 'Column1.' prefix)
        def g(key, alt=''):
            v = rec.get(f'Column1.{key}')
            if v is None:
                v = rec.get(key)
            return v if v is not None else alt

        pid  = str(g('id',   '')).strip()
        name = str(g('name', '')).strip()
        if not pid or not name or name == 'nan':
            skipped += 1
            continue

        cid  = str(g('cid',  '')).strip()
        form = str(g('form', 'Tablet')).strip() or 'Tablet'
        try:    pur = float(g('pur', 0))
        except: pur = 0.0
        try:    mrp = float(g('mrp', 0))
        except: mrp = 0.0
        try:    gst = float(g('gst', 12))
        except: gst = 12.0
        try:    stk = int(float(g('stk', 0)))
        except: stk = 0
        try:    ms  = int(float(g('ms', 10)))
        except: ms  = 10
        bat  = str(g('bat', '')).strip()
        if bat == 'nan': bat = ''
        exp  = str(g('exp', '')).strip()
        if exp == 'nan': exp = ''
        mfr  = str(g('mfr', '')).strip()
        if mfr == 'nan': mfr = ''
        hsn  = str(g('hsn', '')).strip()
        if hsn == 'nan': hsn = ''
        cmp  = str(g('cmp', '')).strip()
        if cmp == 'nan': cmp = ''

        # Ensure category exists
        cat_id = cid if cid else None
        if cat_id:
            exists_cat = conn.execute("SELECT id FROM categories WHERE id=%s", (cat_id,)).fetchone()
            if not exists_cat:
                cat_name = cid_name_map.get(cat_id, f'Category {cat_id[:6]}')
                conn.execute(
                    'INSERT INTO categories(id,name,"desc") VALUES(%s,%s,%s) ON CONFLICT (id) DO NOTHING',
                    (cat_id, cat_name, '')
                )

        try:
            conn.execute("""
                INSERT INTO products
                  (id,name,category,unit,purchase,sale,gst,stock,min_stock,
                   sku,expiry,brand,hsn,"desc",partition)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                  purchase=EXCLUDED.purchase, sale=EXCLUDED.sale, gst=EXCLUDED.gst,
                  stock=EXCLUDED.stock, min_stock=EXCLUDED.min_stock, sku=EXCLUDED.sku,
                  expiry=EXCLUDED.expiry, brand=EXCLUDED.brand, hsn=EXCLUDED.hsn,
                  "desc"=EXCLUDED."desc", partition=EXCLUDED.partition
            """, (pid, name, cat_id, form, pur, mrp, gst, stk, ms,
                  bat, exp, mfr, hsn, cmp, PARTITION_BOTH))
            inserted += 1
        except Exception as e:
            skipped += 1

    conn.commit()
    conn.close()
    print(f"  ✓ Excel medicines seeded: {inserted} inserted, {skipped} skipped")
    print(f"    (partition='both' — visible in Wholesale AND all Retail modes)")


# ─────────────────────────────────────────────────────────────
# SEED DATA  (auto-runs on first launch if DB is empty)
# All seed records get partition='both' — visible in ALL modes.
# ─────────────────────────────────────────────────────────────
def seed_demo_data():
    conn  = get_db()
    # If bills already exist, demo data was already seeded
    count = conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0]
    if count > 0:
        conn.close()
        return  # already has demo billing data

    def mo(n):
        d = date.today()
        m = d.month + n
        y = d.year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        return f"{y}-{m:02d}"

    cats = [
        (uid(), 'Analgesics',             'Pain relievers'),
        (uid(), 'Antibiotics',            'Antibacterial medicines'),
        (uid(), 'Antacids',               'Stomach & digestion'),
        (uid(), 'Antihistamines',         'Allergy medicines'),
        (uid(), 'Vitamins & Supplements', 'Nutritional supplements'),
        (uid(), 'Antidiabetics',          'Diabetes medicines'),
        (uid(), 'Cardiovascular',         'Heart & BP medicines'),
        (uid(), 'Syrups & Liquids',       'Liquid medicines'),
        (uid(), 'Topical',                'Creams, gels & ointments'),
    ]
    for c in cats:
        conn.execute('INSERT INTO categories(id,name,"desc") VALUES(%s,%s,%s) ON CONFLICT (id) DO NOTHING', c)

    # All seed products → partition='both'
    prods = [
        (uid(),'Paracetamol 500mg',  0,'Tablet', 12, 22, 5,200, 50,'B240101',mo(18),'Cipla',         'Paracetamol 500mg',           '30049099'),
        (uid(),'Amoxicillin 250mg',  1,'Capsule',55, 85,12,  8, 20,'SP2024A',mo(10),'Sun Pharma',     'Amoxicillin trihydrate 250mg','30041090'),
        (uid(),'Azithromycin 500mg', 1,'Tablet', 78,120,12, 45, 20,'LU4422', mo(2), 'Lupin',          'Azithromycin 500mg',          '30041090'),
        (uid(),'Cetirizine 10mg',    3,'Tablet', 18, 35, 5,150, 30,'MK0001', mo(24),'Mankind Pharma', 'Cetirizine HCl 10mg',         '30049099'),
        (uid(),'Omeprazole 20mg',    2,'Capsule',30, 55,12,  4, 25,'DR7788', mo(8), "Dr. Reddy's",    'Omeprazole 20mg',             '30049099'),
        (uid(),'Cough Syrup 100ml',  7,'Bottle', 40, 65,12,  5, 15,'PF990X', mo(6), 'Pfizer',         'Dextromethorphan+Guaifenesin','30049039'),
        (uid(),'Metformin 500mg',    5,'Tablet', 20, 42, 5,300,100,'USV0055',mo(16),'USV Ltd',         'Metformin HCl 500mg',         '30049099'),
        (uid(),'Amlodipine 5mg',     6,'Tablet', 22, 38, 5,180, 50,'CF2200', mo(20),'Cadila',          'Amlodipine Besylate 5mg',     '30049099'),
        (uid(),'Vitamin C 500mg',    4,'Tablet', 14, 28, 5,500, 50,'HM8800', mo(30),'Himalaya',        'Ascorbic Acid 500mg',         '29362700'),
        (uid(),'Betadine Cream 10g', 8,'Cream',  28, 48,12, 60, 10,'WM3300', mo(22),'Win Medicare',    'Povidone-Iodine 5%',          '30049039'),
        (uid(),'Eye Drops 5ml',      0,'Drops',  55, 90,12, 12, 10,'AL2023', mo(-2),'Alcon',           'Moxifloxacin 0.5%',           '30049039'),
        (uid(),'Pantoprazole 40mg',  2,'Tablet', 25, 45,12,  0, 30,'SR4411', mo(14),'Serum',           'Pantoprazole Sodium 40mg',    '30049099'),
    ]
    for p in prods:
        pid, name, ci, unit, pur, sal, gst, stk, mns, sku, exp, brand, desc_val, hsn = p
        cat_id = cats[ci][0]
        conn.execute("""
            INSERT INTO products
              (id,name,category,unit,purchase,sale,gst,stock,min_stock,sku,expiry,brand,"desc",hsn,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (pid, name, cat_id, unit, pur, sal, gst, stk, mns, sku, exp, brand, desc_val, hsn,
              PARTITION_BOTH))

    # Demo bills — tagged retail, partition_both handled by bill_store_type
    patients = ['Ramesh Kumar','Priya Sharma','Anil Patel','Sunita Rao',
                'Vijay Singh','Meena Joshi','Deepak Nair']
    doctors  = ['Dr. Mehta','Dr. Singh','Dr. Verma','Dr. Pillai','Dr. Khan']
    modes    = ['Cash','UPI','Cash','Card','UPI','Cash','Insurance','Cash']

    for i in range(14):
        bill_date = (date.today() - timedelta(days=i % 9)).isoformat()
        p1 = prods[i % len(prods)]
        p2 = prods[(i + 3) % len(prods)]

        def make_it(p, qty, disc):
            pur, sal, gst = p[4], p[5], p[6]
            lt  = qty * sal
            da  = lt * disc / 100
            tax = lt - da
            ga  = round(tax * gst / 100, 2)
            ltt = round(tax + ga, 2)
            return (uid(), p[0], p[1], cats[p[2]][0], p[3],
                    qty, sal, disc, gst, ga, ltt)

        items = [make_it(p1, (i%5)+2, 0), make_it(p2, (i%3)+1, 5 if i%3==0 else 0)]
        sub   = sum(it[5]*it[6] for it in items)
        td    = sum(it[5]*it[6]*it[7]/100 for it in items)
        tg    = sum(it[9] for it in items)
        raw   = sub - td + tg
        grand = round(raw)
        bid   = uid()
        bn    = str(i + 1).zfill(4)
        # Seed 12 retail + 2 wholesale demo bills
        b_type = 'wholesale' if i >= 12 else 'retail'
        conn.execute("""
            INSERT INTO bills
              (id,bill_no,date,customer,phone,doctor,payment_mode,
               subtotal,total_discount,total_gst,round_off,grand_total,bill_store_type)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (bid, bn, bill_date,
              patients[i % len(patients)], '', doctors[i % len(doctors)],
              modes[i % len(modes)],
              round(sub,2), round(td,2), round(tg,2), round(grand-raw,2), grand, b_type))
        for it in items:
            conn.execute("""
                INSERT INTO bill_items
                  (id,bill_id,product_id,name,category,unit,qty,unit_price,
                   discount,gst_rate,gst_amt,line_total)
                VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  bill_id=EXCLUDED.bill_id, product_id=EXCLUDED.product_id,
                  name=EXCLUDED.name, category=EXCLUDED.category, unit=EXCLUDED.unit,
                  qty=EXCLUDED.qty, unit_price=EXCLUDED.unit_price,
                  discount=EXCLUDED.discount, gst_rate=EXCLUDED.gst_rate,
                  gst_amt=EXCLUDED.gst_amt, line_total=EXCLUDED.line_total
            """, (it[0], bid) + it[1:])

    # Skip global settings seed — per-user settings are created on registration

    # Demo stock-ins → partition='both'
    for i in range(6):
        p = prods[i * 2 % len(prods)]
        conn.execute("""
            INSERT INTO stock_ins
              (id,date,product_id,product_name,qty,price,batch,expiry,supplier,invoice_no,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (uid(),
              (date.today()-timedelta(days=i+1)).isoformat(),
              p[0], p[1], 50+i*10, p[4], f"B{1000+i}", mo(12+i),
              'Main Distributor', f"INV-{100+i}",
              PARTITION_BOTH))

    # Demo wholesale credits → partition='both'
    c_shops = [
        ('Ramesh Medical Store',  'Ramesh Kumar',    '9876543210'),
        ('Priya Pharma Traders',  'Priya Sharma',    '9823456780'),
        ('Anil Drug House',       'Anil Patel',      '9912345678'),
        ('Sunita Medicals',       'Sunita Rao',      '9988776655'),
        ('Vijay Health Store',    'Vijay Singh',     '9765432109'),
        ('Meena Pharmaceuticals', 'Meena Joshi',     '9654321098'),
        ('Deepak Medical Agency', 'Deepak Nair',     '9543210987'),
        ('Kumar Drug Centre',     'Suresh Kumar',    '9432109876'),
        ('Patel Pharma Dist.',    'Rakesh Patel',    '9321098765'),
        ('Singh Medicals',        'Harpreet Singh',  '9210987654'),
        ('Jain Medical Traders',  'Abhay Jain',      '9109876543'),
        ('Sharma Drug House',     'Mohan Sharma',    '9098765432'),
    ]
    c_items   = ['Paracetamol 500mg x100','Amoxicillin 250mg x50','Azithromycin 500mg x30',
                 'Cetirizine 10mg x200','Omeprazole 20mg x80','Metformin 500mg x150',
                 'Vitamin C 500mg x100','Cough Syrup 100ml x20','Amlodipine 5mg x120',
                 'Betadine Cream 10g x40','Eye Drops 5ml x60','Pantoprazole 40mg x90']
    c_methods = ['UPI','NEFT','Cash','Credit/Debit Card','UPI','NEFT',
                 'Cash','UPI','NEFT','Cash','Credit/Debit Card','UPI']
    c_status  = ['Pending','Cleared','Pending','Pending','Cleared','Pending',
                 'Cleared','Pending','Pending','Cleared','Pending','Cleared']
    c_amounts = [1850,3200,2750,4100,1500,5600,2300,3800,2100,4900,1700,3500]
    c_days    = [2,5,8,12,18,22,28,35,45,60,75,85]
    for i, cs in enumerate(c_shops):
        conn.execute("""
            INSERT INTO credits
              (id,date,shop_name,shopkeeper_name,phone,for_item,amount,method,status,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (uid(), (date.today()-timedelta(days=c_days[i])).isoformat(),
              cs[0], cs[1], cs[2], c_items[i], c_amounts[i], c_methods[i], c_status[i],
              PARTITION_BOTH))

    # Demo retail shop-credits → partition='both'
    suppliers = [
        ('Apex Pharma Dist.',   'WHL-001','Rajesh Gupta',  12500,10000),
        ('MedLine Wholesale',   'WHL-002','Sanjay Mehta',   8200, 8200),
        ('BharatMed Traders',   'WHL-003','Vikram Shah',   15000, 8000),
        ('Sunrise Drug House',  'WHL-004','Pooja Reddy',    6800, 6800),
        ('National Pharma Co.', 'WHL-005','Arvind Kumar',  22000,15000),
        ('HealthFirst Dist.',   'WHL-006','Suresh Nair',    9400, 5000),
        ('Prime Med Supply',    'WHL-007','Deepa Iyer',    11200, 7000),
        ('City Drug Traders',   'WHL-008','Ravi Sharma',    7600, 5000),
        ('Lifeline Wholesale',  'WHL-009','Anita Patel',   18500,12000),
        ('GreenMed Dist.',      'WHL-010','Kartik Joshi',   5300, 3000),
    ]
    s_methods = ['UPI','NEFT','Cash','UPI','NEFT','Cash','Credit/Debit Card','UPI','NEFT','Cash']
    s_days    = [3, 8,12,20,25,32,40,50,62,75]
    for i, sup in enumerate(suppliers):
        pending = round(sup[3] - sup[4], 2)
        d_back  = (date.today()-timedelta(days=s_days[i])).isoformat()
        conn.execute("""
            INSERT INTO shop_credits
              (id,supplier_id,supplier_name,owner_name,total_purchase,paid,
               payment_mode,pending,last_purchase_date,bill_date,status,partition)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id) DO NOTHING
        """, (uid(), sup[1], sup[0], sup[2], sup[3], sup[4],
              s_methods[i], pending, d_back, d_back,
              'Cleared' if pending <= 0 else 'Pending',
              PARTITION_BOTH))

    conn.commit()
    conn.close()
    print("  ✓ Demo data seeded (12 medicines, 14 bills, credits, shop-credits)")
    print("  ✓ All seed records tagged partition='both' (visible in all pharmacy modes)")


# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    # Pre-warm the connection pool BEFORE handling any requests.
    # Without this, the first 4 requests each pay a 200ms TCP handshake.
    print("  ⚡ Pre-warming connection pool…")
    try:
        _get_pool()
        print("  ✓ Connection pool ready (4 warm connections to Supabase)")
    except Exception as e:
        print(f"  ⚠ Pool pre-warm failed: {e} — will retry on first request")

    init_db()
    seed_excel_medicines()
    seed_demo_data()
    print("=" * 62)
    print("  PharmaCare Pro — Flask + PostgreSQL Backend (JWT Auth Edition)")
    print("  Open in browser :  http://localhost:5000")
    print("  Database       :  PostgreSQL (set DATABASE_URL env var)")
    print()
    print("  ── Partition Architecture ─────────────────────────────")
    print("  Settings → Pharmacy Type → determines active partition")
    print()
    print("  Wholesale Pharma   → 'wholesale' partition")
    print("  Retail Pharmacy  ┐")
    print("  Hospital Pharmacy├→ 'retail'    partition")
    print("  Medical Store    │")
    print("  Ayurvedic Store  ┘")
    print()
    print("  partition='both'   → seed/migrated data (all modes)")
    print()
    print("  ── Sub-databases inside pharmacare.db ─────────────────")
    print("  Medicine DB      : products, stock_ins   (partitioned)")
    print("  Sales History DB : bills, bill_items     (partitioned)")
    print("  Credit DB WS     : credits               (wholesale)")
    print("  Credit DB RT     : shop_credits          (retail)")
    print("  ──────────────────────────────────────────────────────")
    print()
    print("  ── Migrate from old localStorage app ─────────────────")
    print("  1. Open OLD app in browser (file:// version)")
    print("  2. Open DevTools Console  (F12 → Console tab)")
    print("  3. Run:  copy(localStorage.getItem('pharmacare_v2'))")
    print("  4. Paste into a file called: state.json")
    print("  5. Run: python migrate.py")
    print("  (Migrated data will be tagged partition='both')")
    print("  ─────────────────────────────────────────────────────")
    print()
    app.run(debug=True, port=5000)
